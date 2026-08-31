#!/usr/bin/env python3
"""Collect auditable evidence for chemical inquiry research without using an LLM.

The collector reads BD/CAS rows from an XLSX or CSV file, queries structured
public sources, optionally queries Tavily or Brave Search, and writes both full normalized
archives and a compact Markdown evidence pack suitable for a later LLM summary.

No OpenAI API or other language-model API is called by this program.
"""

from __future__ import annotations

import argparse
import concurrent.futures
import csv
import dataclasses
import datetime as dt
import hashlib
import html
import json
import math
import os
import re
import sys
import threading
import time
import urllib.parse
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook


SCRIPT_VERSION = "0.4.1"
CACHE_SCHEMA_VERSION = "4"
RESUME_SCHEMA_VERSION = 1
RESUME_STATE_FILENAME = "resume_state.json"
CHECKPOINT_DIRNAME = "checkpoints"
DEFAULT_MCE_MAX_PRODUCTS = 0
PATENT_VERIFICATION_MODES = {
    "off": {
        "scan_pages": 0,
        "final_patents": 0,
        "search_queries": 0,
        "label": "关闭",
    },
    "balanced": {
        "scan_pages": 5,
        "final_patents": 3,
        "search_queries": 1,
        "label": "平衡",
    },
    "deep": {
        "scan_pages": 10,
        "final_patents": 5,
        "search_queries": 3,
        "label": "深度",
    },
}
DEFAULT_SOURCES = (
    "pubchem",
    "googlepatents",
    "chembl",
    "europepmc",
    "pubmed",
    "clinicaltrials",
    "tavily",
)
SUPPORTED_SOURCES = DEFAULT_SOURCES + (
    "mce",
    "brave",
)
CAS_RE = re.compile(r"^\d{2,7}-\d{2}-\d$")
PATENT_RE = re.compile(
    r"\b(?:WO|EP|US|CN|JP|KR|CA|AU|IN)[-\s/]?\d{4,}(?:[-\s/]?[A-Z]\d?)?\b",
    re.IGNORECASE,
)
URL_RE = re.compile(r"https?://[^\s<>\]\[\"']+")
YEAR_RE = re.compile(r"\b(?:19|20)\d{2}\b")


@dataclasses.dataclass(frozen=True)
class ProductInput:
    row_number: int
    bd: str
    cas: str


@dataclasses.dataclass
class EvidenceItem:
    bd: str
    cas: str
    source: str
    source_type: str
    title: str
    url: str
    snippet: str = ""
    published_date: str = ""
    query: str = ""
    identifiers: str = ""
    relevance_score: int = 0
    match_basis: str = "structured_source"
    direct_identifier_match: bool = False
    evidence_scope: str = "direct"


class CollectionError(RuntimeError):
    pass


def utc_now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()


def clean_text(value: Any, limit: int | None = None) -> str:
    if value is None:
        return ""
    text = str(value).replace("\ufffd", " ")
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if limit and len(text) > limit:
        return text[: max(0, limit - 1)].rstrip() + "…"
    return text


def xml_text(node: ET.Element | None) -> str:
    if node is None:
        return ""
    return clean_text("".join(node.itertext()))


def cas_is_valid(cas: str) -> bool:
    if not CAS_RE.fullmatch(cas):
        return False
    digits = cas.replace("-", "")
    body, check = digits[:-1], int(digits[-1])
    total = sum(int(ch) * multiplier for multiplier, ch in enumerate(reversed(body), start=1))
    return total % 10 == check


def mce_allowed_product_rows(products: list[ProductInput], optional_cap: int) -> set[int]:
    """Return every product by default; a positive cap is an explicit test-run choice."""
    selected = products[:optional_cap] if optional_cap > 0 else products
    return {product.row_number for product in selected}


def product_checkpoint_key(product: ProductInput) -> str:
    return f"{product.row_number}:{product.cas}"


def checkpoint_path(output_dir: Path, product: ProductInput) -> Path:
    filename = f"{product.row_number:06d}__{safe_filename(product.cas)}.json"
    return output_dir / CHECKPOINT_DIRNAME / filename


def write_json_atomic(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def resume_collection_config(args: argparse.Namespace) -> dict[str, Any]:
    """Persist result-affecting settings, never API keys."""
    return {
        "sources": sorted(args.sources),
        "cache_dir": str(args.cache_dir.resolve()),
        "cache_ttl_days": args.cache_ttl_days,
        "literature_limit": args.literature_limit,
        "patent_verification": args.patent_verification,
        "patent_pages": args.patent_pages,
        "patent_candidate_cap": args.patent_candidate_cap,
        "patent_snippet_chars": args.patent_snippet_chars,
        "web_results": args.web_results,
        "mce_max_products": args.mce_max_products,
        "include_recent_events": args.include_recent_events,
        "pack_items": args.pack_items,
        "snippet_chars": args.snippet_chars,
        "ncbi_email": args.ncbi_email,
        "user_agent": args.user_agent,
    }


def create_resume_state(
    products: list[ProductInput], args: argparse.Namespace, started_at: str
) -> dict[str, Any]:
    return {
        "schema_version": RESUME_SCHEMA_VERSION,
        "script_version": SCRIPT_VERSION,
        "status": "running",
        "created_at": started_at,
        "updated_at": started_at,
        "last_resumed_at": None,
        "resume_count": 0,
        "input_mode": "single_cas" if args.cas else "file",
        "input_file": str(args.input.resolve()) if args.input else "",
        "input_cas": args.cas or "",
        "products": [dataclasses.asdict(product) for product in products],
        "collection_config": resume_collection_config(args),
        "total_products": len(products),
        "completed_products": 0,
        "remaining_products": len(products),
        "completed_product_keys": [],
        "final_outputs_ready": False,
    }


def load_resume_state(resume_dir: Path) -> dict[str, Any]:
    path = resume_dir / RESUME_STATE_FILENAME
    if not path.exists():
        raise CollectionError(f"续跑目录缺少 {RESUME_STATE_FILENAME}：{resume_dir}")
    try:
        state = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise CollectionError(f"无法读取续跑状态：{exc}") from exc
    if not isinstance(state, dict) or state.get("schema_version") != RESUME_SCHEMA_VERSION:
        raise CollectionError("续跑状态版本不兼容，请保留旧目录并新建任务")
    if state.get("status") == "complete":
        raise CollectionError("该任务已完成，无需断点续跑")
    return state


def apply_resume_config(args: argparse.Namespace, state: dict[str, Any]) -> list[ProductInput]:
    config = state.get("collection_config") or {}
    for name in (
        "cache_ttl_days",
        "literature_limit",
        "patent_verification",
        "patent_pages",
        "patent_candidate_cap",
        "patent_snippet_chars",
        "web_results",
        "mce_max_products",
        "include_recent_events",
        "pack_items",
        "snippet_chars",
        "ncbi_email",
        "user_agent",
    ):
        if name in config:
            setattr(args, name, config[name])
    args.sources = set(config.get("sources") or [])
    args.cache_dir = Path(str(config.get("cache_dir") or args.cache_dir))
    args.output_dir = args.resume_dir
    args.input = Path(state["input_file"]) if state.get("input_file") else None
    args.cas = str(state.get("input_cas") or "")
    products: list[ProductInput] = []
    for value in state.get("products") or []:
        products.append(
            ProductInput(
                row_number=int(value["row_number"]),
                bd=clean_text(value.get("bd")),
                cas=clean_text(value["cas"]),
            )
        )
    if not products:
        raise CollectionError("续跑状态中没有产品列表")
    return products


def load_checkpoint_records(
    output_dir: Path, products: list[ProductInput]
) -> dict[int, dict[str, Any]]:
    records: dict[int, dict[str, Any]] = {}
    for product in products:
        path = checkpoint_path(output_dir, product)
        if not path.exists():
            continue
        try:
            record = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        input_value = record.get("input") or {}
        if (
            int(input_value.get("row_number") or -1) != product.row_number
            or clean_text(input_value.get("cas")) != product.cas
        ):
            continue
        records[product.row_number] = record
    return records


def update_resume_progress(
    output_dir: Path,
    state: dict[str, Any],
    products: list[ProductInput],
    records_by_row: dict[int, dict[str, Any]],
    *,
    status: str | None = None,
    final_outputs_ready: bool | None = None,
) -> None:
    completed_keys = [
        product_checkpoint_key(product)
        for product in products
        if product.row_number in records_by_row
    ]
    state["completed_product_keys"] = completed_keys
    state["completed_products"] = len(completed_keys)
    state["remaining_products"] = len(products) - len(completed_keys)
    state["updated_at"] = utc_now_iso()
    if status is not None:
        state["status"] = status
    if final_outputs_ready is not None:
        state["final_outputs_ready"] = final_outputs_ready
    write_json_atomic(output_dir / RESUME_STATE_FILENAME, state)


def estimate_tokens(text: str) -> int:
    """Conservative local estimate; actual tokenizer/account usage can differ."""
    cjk = sum(1 for ch in text if "\u2e80" <= ch <= "\u9fff")
    other = len(text) - cjk
    return cjk + math.ceil(other / 4)


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s_\-]+", "", clean_text(value).casefold())


def read_products(path: Path, sheet_name: str | None, bd_column: str, cas_column: str) -> list[ProductInput]:
    suffix = path.suffix.lower()
    rows: list[list[Any]]
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise CollectionError(f"工作表不存在：{sheet_name}；可用工作表：{', '.join(workbook.sheetnames)}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    else:
        raise CollectionError("输入文件必须是 .xlsx、.xlsm 或 .csv")

    if not rows:
        raise CollectionError("输入文件没有数据")

    wanted_bd = normalize_header(bd_column)
    wanted_cas = normalize_header(cas_column)
    header_index = None
    bd_index = None
    cas_index = None
    for idx, row in enumerate(rows[:20]):
        normalized = [normalize_header(cell) for cell in row]
        if wanted_cas in normalized:
            header_index = idx
            cas_index = normalized.index(wanted_cas)
            bd_index = normalized.index(wanted_bd) if wanted_bd in normalized else None
            break
    if header_index is None or cas_index is None:
        raise CollectionError(f"前20行未找到 CAS 列（期望列名：{cas_column}）")

    products: list[ProductInput] = []
    seen: set[tuple[str, str]] = set()
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        cas = clean_text(row[cas_index] if cas_index < len(row) else "")
        bd = clean_text(row[bd_index] if bd_index is not None and bd_index < len(row) else "")
        if not cas:
            continue
        key = (bd, cas)
        if key in seen:
            continue
        seen.add(key)
        products.append(ProductInput(row_number=row_number, bd=bd, cas=cas))
    if not products:
        raise CollectionError("CAS 列中没有可处理的记录")
    return products


class HttpClient:
    def __init__(
        self,
        cache_dir: Path,
        ttl_days: float,
        timeout: float,
        retries: int,
        offline: bool,
        user_agent: str,
        ncbi_api_key: str | None,
    ) -> None:
        self.cache_dir = cache_dir
        self.ttl_seconds = ttl_days * 86400
        self.timeout = timeout
        self.retries = retries
        self.offline = offline
        self.user_agent = user_agent
        self.ncbi_api_key = ncbi_api_key
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self._thread_local = threading.local()
        self._locks_guard = threading.Lock()
        self._host_locks: dict[str, threading.Lock] = {}
        self._host_last_request: dict[str, float] = {}

    def _session(self) -> requests.Session:
        session = getattr(self._thread_local, "session", None)
        if session is None:
            session = requests.Session()
            session.headers.update({"User-Agent": self.user_agent, "Accept": "application/json, text/xml;q=0.9, */*;q=0.8"})
            self._thread_local.session = session
        return session

    def _interval_for(self, host: str) -> float:
        if "pubchem.ncbi.nlm.nih.gov" in host:
            return 0.22
        if "eutils.ncbi.nlm.nih.gov" in host:
            return 0.11 if self.ncbi_api_key else 0.36
        if "clinicaltrials.gov" in host:
            return 0.18
        if "patents.google.com" in host:
            return 0.60
        if "api.tavily.com" in host:
            return 0.50
        return 0.12

    def _rate_limit(self, host: str) -> None:
        with self._locks_guard:
            lock = self._host_locks.setdefault(host, threading.Lock())
        with lock:
            now = time.monotonic()
            wait = self._interval_for(host) - (now - self._host_last_request.get(host, 0.0))
            if wait > 0:
                time.sleep(wait)
            self._host_last_request[host] = time.monotonic()

    def _prepared_url(self, url: str, params: dict[str, Any] | None) -> str:
        return requests.Request("GET", url, params=params).prepare().url or url

    def _cache_path(self, prepared_url: str) -> Path:
        digest = hashlib.sha256(f"{CACHE_SCHEMA_VERSION}\n{prepared_url}".encode("utf-8")).hexdigest()
        return self.cache_dir / digest[:2] / f"{digest}.json"

    def get_text(
        self,
        url: str,
        params: dict[str, Any] | None = None,
        headers: dict[str, str] | None = None,
        allow_not_found: bool = False,
    ) -> tuple[str | None, str, bool]:
        prepared_url = self._prepared_url(url, params)
        cache_path = self._cache_path(prepared_url)
        if cache_path.exists():
            try:
                cached = json.loads(cache_path.read_text(encoding="utf-8"))
                fetched = dt.datetime.fromisoformat(cached["fetched_at"])
                age = (dt.datetime.now(dt.timezone.utc) - fetched).total_seconds()
                if self.offline or age <= self.ttl_seconds:
                    return cached.get("body"), prepared_url, True
            except (ValueError, KeyError, OSError):
                pass
        if self.offline:
            raise CollectionError(f"离线模式下没有缓存：{prepared_url}")

        host = urllib.parse.urlsplit(prepared_url).hostname or ""
        last_error: Exception | None = None
        for attempt in range(self.retries + 1):
            self._rate_limit(host)
            try:
                response = self._session().get(prepared_url, headers=headers, timeout=self.timeout)
                if response.status_code == 404 and allow_not_found:
                    payload = {
                        "url": prepared_url,
                        "status": 404,
                        "content_type": response.headers.get("Content-Type", ""),
                        "fetched_at": utc_now_iso(),
                        "body": None,
                    }
                    cache_path.parent.mkdir(parents=True, exist_ok=True)
                    cache_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
                    return None, prepared_url, False
                if response.status_code in {429, 500, 502, 503, 504}:
                    retry_after = response.headers.get("Retry-After")
                    delay = float(retry_after) if retry_after and retry_after.isdigit() else min(8.0, 0.8 * (2**attempt))
                    if attempt < self.retries:
                        time.sleep(delay)
                        continue
                response.raise_for_status()
                # The APIs and Google Patents pages used here are UTF-8. Some
                # servers omit a charset, causing requests to assume Latin-1
                # and produce mojibake in chemical procedures.
                try:
                    body = response.content.decode("utf-8-sig")
                except UnicodeDecodeError:
                    body = response.text
                payload = {
                    "url": prepared_url,
                    "status": response.status_code,
                    "content_type": response.headers.get("Content-Type", ""),
                    "fetched_at": utc_now_iso(),
                    "body": body,
                }
                cache_path.parent.mkdir(parents=True, exist_ok=True)
                cache_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
                return body, prepared_url, False
            except (requests.RequestException, ValueError) as exc:
                last_error = exc
                if attempt < self.retries:
                    time.sleep(min(8.0, 0.8 * (2**attempt)))
        raise CollectionError(f"请求失败：{prepared_url}；{last_error}")

    def get_json(
        self,
        url: str,
        params: dict[str, Any] | None = None,
        headers: dict[str, str] | None = None,
        allow_not_found: bool = False,
    ) -> tuple[Any | None, str, bool]:
        body, prepared_url, cached = self.get_text(url, params, headers, allow_not_found)
        if body is None:
            return None, prepared_url, cached
        try:
            return json.loads(body), prepared_url, cached
        except json.JSONDecodeError as exc:
            raise CollectionError(f"接口返回的不是有效 JSON：{prepared_url}；{exc}") from exc

    def post_json(
        self,
        url: str,
        payload: dict[str, Any],
        headers: dict[str, str] | None = None,
    ) -> tuple[Any, str, bool]:
        """POST JSON with a body-aware cache; authorization headers are never cached."""
        canonical_payload = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        cache_path = self._cache_path(f"POST\n{url}\n{canonical_payload}")
        if cache_path.exists():
            try:
                cached = json.loads(cache_path.read_text(encoding="utf-8"))
                fetched = dt.datetime.fromisoformat(cached["fetched_at"])
                age = (dt.datetime.now(dt.timezone.utc) - fetched).total_seconds()
                if self.offline or age <= self.ttl_seconds:
                    return json.loads(cached["body"]), url, True
            except (ValueError, KeyError, OSError, json.JSONDecodeError):
                pass
        if self.offline:
            raise CollectionError(f"离线模式下没有缓存：POST {url}")

        host = urllib.parse.urlsplit(url).hostname or ""
        last_error: Exception | None = None
        request_headers = {"Content-Type": "application/json", **(headers or {})}
        for attempt in range(self.retries + 1):
            self._rate_limit(host)
            try:
                response = self._session().post(url, headers=request_headers, json=payload, timeout=self.timeout)
                if response.status_code in {429, 500, 502, 503, 504}:
                    retry_after = response.headers.get("Retry-After")
                    delay = float(retry_after) if retry_after and retry_after.isdigit() else min(8.0, 0.8 * (2**attempt))
                    if attempt < self.retries:
                        time.sleep(delay)
                        continue
                response.raise_for_status()
                try:
                    body = response.content.decode("utf-8-sig")
                except UnicodeDecodeError:
                    body = response.text
                data = json.loads(body)
                cache_record = {
                    "url": url,
                    "method": "POST",
                    "status": response.status_code,
                    "content_type": response.headers.get("Content-Type", ""),
                    "fetched_at": utc_now_iso(),
                    "body": body,
                }
                cache_path.parent.mkdir(parents=True, exist_ok=True)
                cache_path.write_text(json.dumps(cache_record, ensure_ascii=False), encoding="utf-8")
                return data, url, False
            except (requests.RequestException, ValueError, json.JSONDecodeError) as exc:
                last_error = exc
                if attempt < self.retries:
                    time.sleep(min(8.0, 0.8 * (2**attempt)))
        raise CollectionError(f"请求失败：POST {url}；{last_error}")


def flatten_json_strings(value: Any, output: list[str], max_items: int = 5000) -> None:
    if len(output) >= max_items:
        return
    if isinstance(value, str):
        text = clean_text(value)
        if text:
            output.append(text)
    elif isinstance(value, dict):
        for child in value.values():
            flatten_json_strings(child, output, max_items)
    elif isinstance(value, list):
        for child in value:
            flatten_json_strings(child, output, max_items)


def extract_year(value: str) -> str:
    match = YEAR_RE.search(value or "")
    return match.group(0) if match else ""


def patent_url(identifier: str) -> str:
    normalized = re.sub(r"[^A-Z0-9]", "", identifier.upper())
    return f"https://patents.google.com/patent/{urllib.parse.quote(normalized)}/en"


def collect_pubchem(product: ProductInput, client: HttpClient) -> tuple[dict[str, Any], list[EvidenceItem]]:
    cas_path = urllib.parse.quote(product.cas, safe="")
    cids_data, cids_url, _ = client.get_json(
        f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{cas_path}/cids/JSON",
        allow_not_found=True,
    )
    cids = (((cids_data or {}).get("IdentifierList") or {}).get("CID") or [])
    if not cids:
        return {"query_url": cids_url, "found": False, "cids": []}, []
    cid = int(cids[0])
    property_names = ",".join(
        [
            "Title",
            "IUPACName",
            "MolecularFormula",
            "MolecularWeight",
            "SMILES",
            "InChI",
            "InChIKey",
            "ExactMass",
            "MonoisotopicMass",
        ]
    )
    props_data, props_url, _ = client.get_json(
        f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/property/{property_names}/JSON"
    )
    properties = (((props_data or {}).get("PropertyTable") or {}).get("Properties") or [{}])[0]
    synonyms_data, synonyms_url, _ = client.get_json(
        f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/synonyms/JSON",
        allow_not_found=True,
    )
    synonyms = (
        ((((synonyms_data or {}).get("InformationList") or {}).get("Information") or [{}])[0]).get("Synonym")
        or []
    )
    synonyms = [clean_text(item) for item in synonyms if clean_text(item)][:100]

    sections: dict[str, Any] = {}
    patent_ids: set[str] = set()
    for heading in ("Depositor-Supplied Patent Identifiers", "Drug and Medication Information"):
        section_data, section_url, _ = client.get_json(
            f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON",
            params={"heading": heading},
            allow_not_found=True,
        )
        if not section_data:
            continue
        strings: list[str] = []
        flatten_json_strings(section_data, strings)
        combined = clean_text(" | ".join(strings), 12000)
        for match in PATENT_RE.findall(combined):
            patent_ids.add(clean_text(match).upper())
        sections[heading] = {"url": section_url, "text": combined, "data": section_data}

    patent_xrefs_text, patent_xrefs_url, _ = client.get_text(
        f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/xrefs/PatentID/TXT",
        allow_not_found=True,
    )
    for line in (patent_xrefs_text or "").splitlines()[:5000]:
        identifier = clean_text(line).upper()
        if identifier and len(identifier) <= 80:
            patent_ids.add(identifier)

    title = clean_text(properties.get("Title") or properties.get("IUPACName") or product.cas)
    compound_url = f"https://pubchem.ncbi.nlm.nih.gov/compound/{cid}"
    identity_bits = [
        f"CID {cid}",
        title,
        clean_text(properties.get("MolecularFormula")),
        f"MW {clean_text(properties.get('MolecularWeight'))}" if properties.get("MolecularWeight") else "",
        f"InChIKey {clean_text(properties.get('InChIKey'))}" if properties.get("InChIKey") else "",
    ]
    evidence = [
        EvidenceItem(
            bd=product.bd,
            cas=product.cas,
            source="PubChem",
            source_type="chemical_identity",
            title=f"PubChem compound record: {title}",
            url=compound_url,
            snippet="; ".join(bit for bit in identity_bits if bit),
            identifiers=f"CID:{cid}",
            relevance_score=100,
            match_basis="structured_cas_query",
            direct_identifier_match=True,
        )
    ]
    for heading, section in sections.items():
        # The patent PUG View section is mostly a generic description and a
        # link to PatentID/TXT. The actual identifiers collected below are
        # much more useful to a summarizer than repeating that boilerplate.
        if section["text"] and "Patent" not in heading:
            evidence.append(
                EvidenceItem(
                    bd=product.bd,
                    cas=product.cas,
                    source="PubChem PUG View",
                    source_type="patent" if "Patent" in heading else "drug_annotation",
                    title=heading,
                    url=section["url"],
                    snippet=clean_text(section["text"], 1400),
                    relevance_score=88 if "Patent" in heading else 84,
                    match_basis="pubchem_compound_section",
                    direct_identifier_match=True,
                )
            )
    sorted_patent_ids = sorted(patent_ids)
    if sorted_patent_ids:
        evidence.append(
            EvidenceItem(
                bd=product.bd,
                cas=product.cas,
                source="PubChem patent cross-references",
                source_type="patent_index",
                title=f"PubChem patent identifiers ({len(sorted_patent_ids)} candidates)",
                url=patent_xrefs_url,
                snippet=clean_text(
                    f"PubChem lists {len(sorted_patent_ids)} depositor-supplied patent identifiers. "
                    "These are index candidates, not proof that a patent uses the compound in a claimed synthesis.",
                    1400,
                ),
                identifiers=";".join(sorted_patent_ids[:20]),
                relevance_score=82,
                match_basis="pubchem_patent_index",
                direct_identifier_match=False,
                evidence_scope="candidate",
            )
        )
    return {
        "found": True,
        "query_url": cids_url,
        "property_url": props_url,
        "synonyms_url": synonyms_url,
        "cids": cids[:10],
        "selected_cid": cid,
        "properties": properties,
        "synonyms": synonyms,
        "sections": sections,
        "patent_xrefs_url": patent_xrefs_url,
        "patent_id_count": len(sorted_patent_ids),
        "patent_ids": sorted_patent_ids[:500],
    }, evidence


def normalize_patent_identifier(identifier: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", clean_text(identifier).upper())


def patent_publication_parts(identifier: str) -> tuple[str, str, str]:
    normalized = normalize_patent_identifier(identifier)
    match = re.match(r"^([A-Z]{2})(\d{4,})([A-Z]\d?)?$", normalized)
    if not match:
        return "", normalized, ""
    return match.group(1), match.group(2), match.group(3) or ""


def patent_publication_root(identifier: str) -> str:
    country, number, _ = patent_publication_parts(identifier)
    return f"{country}{number}" if country else normalize_patent_identifier(identifier)


def patent_identifier_year(identifier: str) -> int:
    _, number, _ = patent_publication_parts(identifier)
    match = re.match(r"((?:19|20)\d{2})", number)
    return int(match.group(1)) if match else 0


def patent_identifier_priority(identifier: str) -> tuple[int, int, int, str]:
    normalized = normalize_patent_identifier(identifier)
    country, _, kind = patent_publication_parts(normalized)
    country_rank = {
        "WO": 0,
        "EP": 1,
        "US": 2,
        "CN": 3,
        "JP": 4,
        "KR": 5,
        "CA": 6,
        "AU": 7,
        "IN": 8,
    }.get(country, 9)
    kind_rank = 0 if kind in {"A", "A1", "A2"} else 1
    return -patent_identifier_year(normalized), country_rank, kind_rank, normalized


def dedupe_patent_identifiers(identifiers: Iterable[str], cap: int = 300) -> list[str]:
    """Collapse publication/grant variants such as CN...A and CN...B before fetching."""
    best: dict[str, str] = {}
    for raw in identifiers:
        identifier = normalize_patent_identifier(raw)
        if not identifier:
            continue
        root = patent_publication_root(identifier)
        existing = best.get(root)
        if existing is None or patent_identifier_priority(identifier) < patent_identifier_priority(existing):
            best[root] = identifier
        if len(best) >= max(1, cap):
            break
    return sorted(best.values(), key=patent_identifier_priority)


def patent_meta_values(page_html: str) -> dict[str, list[str]]:
    values: defaultdict[str, list[str]] = defaultdict(list)
    for tag in re.findall(r"<meta\b[^>]*>", page_html, flags=re.IGNORECASE):
        attrs = {
            key.casefold(): html.unescape(value)
            for key, _, value in re.findall(
                r"([:\w-]+)\s*=\s*([\"'])(.*?)\2",
                tag,
                flags=re.IGNORECASE | re.DOTALL,
            )
        }
        name = clean_text(attrs.get("name") or attrs.get("property")).casefold()
        content = clean_text(attrs.get("content"), 10000)
        if name and content:
            values[name].append(content)
    return dict(values)


def itemprop_values(page_html: str, property_name: str, limit: int = 50) -> list[str]:
    escaped = re.escape(property_name)
    pattern = re.compile(
        rf"<(?P<tag>[A-Za-z0-9]+)\b(?P<attrs>[^>]*\bitemprop\s*=\s*([\"']){escaped}\3[^>]*)>"
        rf"(?P<body>.*?)</(?P=tag)>",
        flags=re.IGNORECASE | re.DOTALL,
    )
    values: list[str] = []
    for match in pattern.finditer(page_html):
        attrs = match.group("attrs")
        attribute_value = re.search(
            r"\b(?:content|datetime)\s*=\s*([\"'])(.*?)\1",
            attrs,
            flags=re.IGNORECASE | re.DOTALL,
        )
        value = clean_text(
            html.unescape(attribute_value.group(2)) if attribute_value else visible_html_text(match.group("body")),
            1000,
        )
        if value and value not in values:
            values.append(value)
        if len(values) >= limit:
            break
    return values


def patent_page_metadata(page_html: str, patent_id: str) -> dict[str, Any]:
    metadata = patent_meta_values(page_html)
    family_match = re.search(
        r"<section\b[^>]*\bitemprop\s*=\s*([\"'])family\1[^>]*>.*?<h2>\s*ID\s*=\s*([^<]+)</h2>",
        page_html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    legal_match = re.search(
        r"\bitemprop\s*=\s*([\"'])legalStatusIfi\1[^>]*>.*?\bitemprop\s*=\s*([\"'])status\2[^>]*>(.*?)</",
        page_html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    publication_dates = itemprop_values(page_html, "publicationDate", limit=5)
    priority_dates = itemprop_values(page_html, "priorityDate", limit=5)
    family_block_match = re.search(
        r"<section\b[^>]*\bitemprop\s*=\s*([\"'])family\1[^>]*>(.*?)</section>",
        page_html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    family_block = family_block_match.group(2) if family_block_match else ""
    return {
        "title": clean_text((metadata.get("dc.title") or [patent_id])[0]),
        "description": clean_text((metadata.get("dc.description") or [""])[0], 2400),
        "dates": [clean_text(value) for value in metadata.get("dc.date", []) if clean_text(value)],
        "publication_date": publication_dates[0] if publication_dates else "",
        "priority_date": priority_dates[0] if priority_dates else "",
        "assignees": itemprop_values(page_html, "assigneeCurrent", limit=10)
        or itemprop_values(page_html, "assigneeOriginal", limit=10),
        "legal_status": clean_text(visible_html_text(legal_match.group(3))) if legal_match else "",
        "family_id": clean_text(family_match.group(2)) if family_match else "",
        "family_publications": itemprop_values(family_block, "representativePublication", limit=50),
    }


def visible_html_text(page_html: str) -> str:
    without_code = re.sub(
        r"<(?:script|style|noscript)\b[^>]*>.*?</(?:script|style|noscript)>",
        " ",
        page_html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    without_tags = re.sub(r"<[^>]+>", " ", without_code)
    return clean_text(html.unescape(without_tags))


def patent_section_texts(page_html: str, description_fallback: str = "") -> dict[str, str]:
    sections: dict[str, str] = {}
    for section_name in ("claims", "description", "abstract"):
        pattern = re.compile(
            rf"<(?P<tag>[A-Za-z0-9]+)\b[^>]*\bitemprop\s*=\s*([\"']){section_name}\2[^>]*>"
            rf"(?P<body>.*?)</(?P=tag)>",
            flags=re.IGNORECASE | re.DOTALL,
        )
        matches = [visible_html_text(match.group("body")) for match in pattern.finditer(page_html)]
        text = clean_text(" ".join(value for value in matches if value), 400000)
        if text:
            sections[section_name] = text
    if description_fallback and "abstract" not in sections:
        sections["abstract"] = description_fallback
    return sections


def matching_contexts(text: str, terms: list[str], context_chars: int = 420, limit: int = 4) -> tuple[str, list[str]]:
    lowered = text.casefold()
    for term in terms:
        term = clean_text(term)
        if len(term) < 4:
            continue
        needle = term.casefold()
        contexts: list[str] = []
        start_at = 0
        while len(contexts) < limit:
            position = lowered.find(needle, start_at)
            if position < 0:
                break
            left = max(0, position - context_chars)
            right = min(len(text), position + len(term) + context_chars)
            contexts.append(clean_text(text[left:right], context_chars * 2 + len(term) + 20))
            start_at = position + len(term)
        if contexts:
            return term, contexts
    return "", []


def distinctive_patent_terms(product: ProductInput, pubchem: dict[str, Any], fallback_name: str) -> list[str]:
    candidates = [
        fallback_name,
        clean_text((pubchem.get("properties") or {}).get("IUPACName")),
        *(pubchem.get("synonyms") or []),
    ]
    terms: list[str] = []
    for raw in candidates:
        term = clean_text(raw)
        if not term or term.casefold() == product.cas.casefold() or CAS_RE.fullmatch(term):
            continue
        if len(term) < 8 or not re.search(r"[A-Za-z]", term):
            continue
        if re.fullmatch(r"(?:CID|SCHEMBL|AKOS|CS)-?[A-Z0-9._-]+", term, flags=re.IGNORECASE):
            continue
        if term.casefold() not in {value.casefold() for value in terms}:
            terms.append(term)
        if len(terms) >= 30:
            break
    return terms


def related_form_query(fallback_name: str) -> str:
    name = clean_text(fallback_name)
    lowered = name.casefold()
    if "boronic acid" in lowered:
        positional = re.fullmatch(
            r"\(6-cyano-(\d+)-(fluoro|chloro|bromo|iodo)pyridin-(\d+)-yl\)boronic acid",
            lowered,
        )
        if positional:
            halogen_position = 8 - int(positional.group(1))
            boron_position = 8 - int(positional.group(3))
            return (
                f'"{halogen_position}-{positional.group(2)}-{boron_position}-'
                '(4,4,5,5-tetramethyl-1,3,2-dioxaborolan-2-yl)picolinonitrile"'
            )
        stem = re.sub(r"\bboronic acid\b", "", name, flags=re.IGNORECASE).strip(" ()-,")
        return f'"{stem}" (pinacol OR boronate OR dioxaborolan)'
    if lowered.endswith(" acid"):
        stem = name[:-5].strip()
        return f'"{stem}" (ester OR salt)'
    return ""


def chemical_signature_tokens(text: str) -> list[str]:
    normalized = clean_text(text).casefold()
    normalized = re.sub(r"picolinonitrile", " pyridine nitrile ", normalized)
    normalized = re.sub(r"([a-z]+)nitrile\b", r"\1 nitrile", normalized)
    normalized = re.sub(r"\b(fluoro|chloro|bromo|iodo|cyano)(?=[a-z])", r"\1 ", normalized)
    normalized = re.sub(r"pyridin(?:e)?(?=boronic|boronate)", "pyridine ", normalized)
    normalized = re.sub(r"\bpyridin(?:e)?\b", " pyridine ", normalized)
    normalized = re.sub(r"\b(?:boronic|boronate|dioxaborolan|boron)\w*\b", " boron ", normalized)
    normalized = re.sub(r"\bcyano\b", " nitrile ", normalized)
    stop = {
        "acid",
        "compound",
        "hydrate",
        "solution",
        "racemate",
        "mixture",
    }
    tokens: list[str] = []
    for token in re.findall(r"[a-z]{4,}", normalized):
        if token in stop or token in tokens:
            continue
        tokens.append(token)
    return tokens[:16]


def chemical_signature_contexts(
    text: str,
    name_terms: list[str],
    context_chars: int = 650,
    limit: int = 3,
) -> tuple[str, list[str], bool]:
    best_term = ""
    best_contexts: list[str] = []
    best_ratio = 0.0
    best_related_form = False
    lowered = text.casefold()
    for name in name_terms[:8]:
        target = chemical_signature_tokens(name)
        if len(target) < 3:
            continue
        anchors = [token for token in target if len(token) >= 5]
        contexts: list[str] = []
        related_form = False
        checked_positions: set[int] = set()
        for anchor in anchors:
            raw_anchors = [anchor]
            if anchor == "nitrile":
                raw_anchors.append("cyano")
            if anchor == "pyridine":
                raw_anchors.extend(["pyridin", "picolino"])
            if anchor == "boron":
                raw_anchors.extend(["boronic", "boronate", "dioxaborolan"])
            for raw_anchor in raw_anchors:
                start_at = 0
                while len(contexts) < limit:
                    position = lowered.find(raw_anchor, start_at)
                    if position < 0:
                        break
                    start_at = position + len(raw_anchor)
                    if any(abs(position - previous) < 100 for previous in checked_positions):
                        continue
                    checked_positions.add(position)
                    left = max(0, position - context_chars)
                    right = min(len(text), position + len(raw_anchor) + context_chars)
                    context = clean_text(text[left:right], context_chars * 2 + 80)
                    context_tokens = set(chemical_signature_tokens(context))
                    overlap = len(set(target).intersection(context_tokens))
                    ratio = overlap / len(set(target))
                    required = max(3, math.ceil(len(set(target)) * 0.7))
                    if overlap >= required:
                        contexts.append(context)
                        best_ratio = max(best_ratio, ratio)
                        if re.search(r"\b(pinacol|dioxaborolan|boronate ester)\b", context, flags=re.IGNORECASE):
                            related_form = True
        if contexts and (not best_contexts or len(contexts) > len(best_contexts)):
            best_term = name
            best_contexts = contexts
            best_related_form = related_form
    return best_term, best_contexts, best_related_form


def extract_patent_ids_from_search_page(page_html: str, cap: int = 100) -> list[str]:
    values: list[str] = []
    for match in re.finditer(r"(?:href\s*=\s*[\"'])?/patent/([A-Z]{2}[A-Z0-9-]{4,})/", page_html, flags=re.IGNORECASE):
        identifier = normalize_patent_identifier(match.group(1))
        if identifier and identifier not in values:
            values.append(identifier)
        if len(values) >= cap:
            break
    return values


def extract_patent_ids_from_search_data(data: Any, cap: int = 100) -> list[str]:
    values: list[str] = []
    clusters = ((data or {}).get("results") or {}).get("cluster") or [] if isinstance(data, dict) else []
    for cluster in clusters:
        if not isinstance(cluster, dict):
            continue
        for result in cluster.get("result") or []:
            if not isinstance(result, dict):
                continue
            patent = result.get("patent") or {}
            identifier = clean_text(patent.get("publication_number")) if isinstance(patent, dict) else ""
            if not identifier:
                identifier = clean_text(result.get("id")).split("/")[1] if "/" in clean_text(result.get("id")) else ""
            identifier = normalize_patent_identifier(identifier)
            if identifier and identifier not in values:
                values.append(identifier)
            if len(values) >= cap:
                return values
    return values


def patent_candidate_order(
    indexed_ids: Iterable[str],
    searched_ids: Iterable[str],
    page_limit: int,
    candidate_cap: int,
) -> list[str]:
    indexed = dedupe_patent_identifiers(indexed_ids, candidate_cap)
    searched = dedupe_patent_identifiers(searched_ids, candidate_cap)
    ordered: list[str] = []

    def add(identifier: str) -> None:
        root = patent_publication_root(identifier)
        if root and all(patent_publication_root(existing) != root for existing in ordered):
            ordered.append(identifier)

    # Exact Google Patents query hits are the strongest cheap pre-filter.
    searched_quota = max(1, math.ceil(page_limit * 0.6))
    for identifier in searched[:searched_quota]:
        add(identifier)
        if len(ordered) >= page_limit:
            return ordered
    # Mix recent publications with jurisdiction and age diversity. This avoids
    # treating "latest" as the only definition of value when an older patent
    # contains the actual synthesis example.
    recent_quota = max(1, math.ceil((page_limit - len(ordered)) * 0.6))
    for identifier in indexed[:recent_quota]:
        add(identifier)
        if len(ordered) >= page_limit:
            return ordered
    seen_countries: set[str] = set()
    for identifier in indexed:
        country, _, _ = patent_publication_parts(identifier)
        if country and country not in seen_countries:
            seen_countries.add(country)
            add(identifier)
        if len(ordered) >= page_limit:
            return ordered
    for identifier in reversed(indexed):
        add(identifier)
        if len(ordered) >= page_limit:
            return ordered
    for identifier in indexed:
        add(identifier)
        if len(ordered) >= page_limit:
            return ordered
    return ordered


def patent_match_assessment(
    product: ProductInput,
    sections: dict[str, str],
    name_terms: list[str],
) -> dict[str, Any]:
    best: dict[str, Any] = {
        "matched_term": "",
        "match_type": "none",
        "section": "",
        "contexts": [],
        "strength": 0,
        "signals": [],
    }
    section_rank = {"claims": 30, "description": 18, "abstract": 8}
    terms = [(product.cas, "cas_exact", 60), *[(term, "name_or_synonym_exact", 45) for term in name_terms]]
    for section_name in ("claims", "description", "abstract"):
        text = sections.get(section_name, "")
        if not text:
            continue
        for term, match_type, base_score in terms:
            matched_term, contexts = matching_contexts(text, [term], context_chars=520, limit=4)
            if not contexts:
                continue
            joined = " ".join(contexts).casefold()
            signals: list[str] = []
            signal_score = 0
            if section_name == "claims" or re.search(r"\bclaim(?:s|ed)?\b", joined):
                signals.append("claim")
                signal_score += 12
            if re.search(r"\b(example|intermediate|preparation|prepared|synthesi[sz]ed|step|yield)\b", joined):
                signals.append("synthesis_example")
                signal_score += 10
            if re.search(r"\b(ic50|ec50|ki|kd|inhibition|inhibitor|assay|potency|activity)\b", joined):
                signals.append("activity")
                signal_score += 10
            strength = min(100, base_score + section_rank.get(section_name, 0) + signal_score)
            if strength > int(best["strength"]):
                best = {
                    "matched_term": matched_term,
                    "match_type": match_type,
                    "section": section_name,
                    "contexts": contexts,
                    "strength": strength,
                    "signals": signals,
                }
        signature_term, signature_contexts, related_form = chemical_signature_contexts(text, name_terms)
        if signature_contexts:
            joined = " ".join(signature_contexts).casefold()
            signals = ["chemical_name_signature"]
            signal_score = 0
            if section_name == "claims" or re.search(r"\bclaim(?:s|ed)?\b", joined):
                signals.append("claim")
                signal_score += 10
            if re.search(r"\b(example|intermediate|preparation|prepared|synthesi[sz]ed|step|yield)\b", joined):
                signals.append("synthesis_example")
                signal_score += 10
            if re.search(r"\b(ic50|ec50|ki|kd|inhibition|inhibitor|assay|potency|activity)\b", joined):
                signals.append("activity")
                signal_score += 8
            match_type = "related_form_signature" if related_form else "chemical_name_signature"
            strength = min(96, (31 if related_form else 38) + section_rank.get(section_name, 0) + signal_score)
            if strength > int(best["strength"]):
                best = {
                    "matched_term": signature_term,
                    "match_type": match_type,
                    "section": section_name,
                    "contexts": signature_contexts,
                    "strength": strength,
                    "signals": signals,
                }
    return best


def patent_record_score(record: dict[str, Any]) -> int:
    score = int(record.get("strength") or 0)
    publication_year = int(extract_year(record.get("publication_date") or "") or 0)
    current_year = dt.date.today().year
    if publication_year >= current_year - 2:
        score += 5
    elif publication_year >= current_year - 5:
        score += 3
    if clean_text(record.get("legal_status")).casefold() in {"active", "pending"}:
        score += 4
    if record.get("assignees"):
        score += 2
    return min(100, score)


def select_verified_patent_records(records: list[dict[str, Any]], limit: int) -> list[dict[str, Any]]:
    if limit <= 0:
        return []
    family_best: dict[str, dict[str, Any]] = {}
    for record in records:
        if record.get("match_type") == "none":
            continue
        family_key = clean_text(record.get("family_id")) or patent_publication_root(record.get("patent_id", ""))
        existing = family_best.get(family_key)
        if existing is None or int(record.get("score") or 0) > int(existing.get("score") or 0):
            family_best[family_key] = record
    candidates = list(family_best.values())
    selected: list[dict[str, Any]] = []

    def add(record: dict[str, Any]) -> None:
        if record not in selected and len(selected) < limit:
            selected.append(record)

    direct = sorted(candidates, key=lambda value: (-int(value.get("score") or 0), value.get("patent_id", "")))
    for record in direct[:2]:
        add(record)
    recent = sorted(
        candidates,
        key=lambda value: (value.get("publication_date") or value.get("priority_date") or "", int(value.get("score") or 0)),
        reverse=True,
    )
    for record in recent[:2]:
        add(record)
    project_value = sorted(
        candidates,
        key=lambda value: (
            "claim" in (value.get("signals") or []),
            "activity" in (value.get("signals") or []),
            bool(value.get("assignees")),
            int(value.get("score") or 0),
        ),
        reverse=True,
    )
    if project_value:
        add(project_value[0])
    for record in direct:
        add(record)
    return selected[:limit]


def collect_google_patents(
    product: ProductInput,
    client: HttpClient,
    pubchem: dict[str, Any],
    fallback_name: str,
    verification_mode: str = "balanced",
    page_limit_override: int = 0,
    candidate_cap: int = 300,
    excerpt_chars: int = 1200,
) -> tuple[dict[str, Any], list[EvidenceItem]]:
    mode = verification_mode if verification_mode in PATENT_VERIFICATION_MODES else "balanced"
    config = PATENT_VERIFICATION_MODES[mode]
    page_limit = max(0, page_limit_override) or int(config["scan_pages"])
    final_limit = int(config["final_patents"])
    patent_ids = [clean_text(item) for item in (pubchem.get("patent_ids") or []) if clean_text(item)]
    deduped_indexed = dedupe_patent_identifiers(patent_ids, candidate_cap)
    if mode == "off" or page_limit <= 0:
        return {
            "mode": mode,
            "skipped": True,
            "reason": "patent full-text verification is disabled",
            "candidate_count": len(patent_ids),
            "deduped_candidate_count": len(deduped_indexed),
            "selected_count": 0,
            "records": [],
            "cache_hits": 0,
            "network_requests_this_run": 0,
        }, []

    # In balanced mode, a short PubChem list already fits within the page cap;
    # querying Google again would add load without changing coverage.
    search_queries = (
        []
        if mode == "balanced" and 0 < len(deduped_indexed) <= page_limit
        else [f'"{product.cas}"']
    )
    if mode == "deep" and fallback_name:
        search_queries.append(f'"{fallback_name}"')
        form_query = related_form_query(fallback_name)
        if form_query:
            search_queries.append(form_query)
    search_queries = search_queries[: int(config["search_queries"])]
    searched_ids: list[str] = []
    search_records: list[dict[str, Any]] = []
    cache_hits = 0
    network_requests = 0
    for query in search_queries:
        try:
            search_data, search_url, cached = client.get_json(
                "https://patents.google.com/xhr/query",
                params={"url": "q=" + urllib.parse.quote(query) + "&sort=new"},
                allow_not_found=True,
            )
        except Exception as exc:
            network_requests += 1
            search_records.append(
                {
                    "query": query,
                    "url": "https://patents.google.com/xhr/query",
                    "error": f"{type(exc).__name__}: {exc}",
                    "candidate_ids": [],
                }
            )
            # Candidate search is optional. A service-level failure usually
            # affects all sibling queries, so do not hammer the endpoint.
            break
        cache_hits += int(cached)
        network_requests += int(not cached)
        ids = extract_patent_ids_from_search_data(search_data, cap=100)
        searched_ids.extend(ids)
        search_records.append(
            {
                "query": query,
                "url": search_url,
                "reported_result_count": int((((search_data or {}).get("results") or {}).get("total_num_results") or 0)),
                "candidate_ids": ids,
            }
        )

    selected_ids = patent_candidate_order(patent_ids, searched_ids, page_limit, candidate_cap)
    records: list[dict[str, Any]] = []
    name_terms = distinctive_patent_terms(product, pubchem, fallback_name)
    high_confidence_hits = 0
    for patent_id in selected_ids:
        page_url = patent_url(patent_id)
        try:
            page_html, final_url, cached = client.get_text(page_url, allow_not_found=True)
        except Exception as exc:
            network_requests += 1
            records.append(
                {
                    "patent_id": patent_id,
                    "url": page_url,
                    "found": False,
                    "error": f"{type(exc).__name__}: {exc}",
                }
            )
            continue
        cache_hits += int(cached)
        network_requests += int(not cached)
        if not page_html:
            records.append({"patent_id": patent_id, "url": final_url, "found": False})
            continue
        metadata = patent_page_metadata(page_html, patent_id)
        sections = patent_section_texts(page_html, metadata["description"])
        assessment = patent_match_assessment(product, sections, name_terms)
        record = {
            "patent_id": patent_id,
            "url": final_url,
            "found": True,
            **metadata,
            **assessment,
        }
        record["score"] = patent_record_score(record)
        records.append(record)
        if assessment["match_type"] != "none" and int(assessment["strength"]) >= 85:
            high_confidence_hits += 1
        if high_confidence_hits >= 2:
            break

    verified = select_verified_patent_records(records, final_limit)
    verified_ids = {record["patent_id"] for record in verified}
    evidence: list[EvidenceItem] = []
    # Preserve unmatched fetched pages as indirect audit candidates, but only
    # selected verified records are eligible for the LLM evidence pack.
    archived = verified + [record for record in records if record.get("found") and record.get("patent_id") not in verified_ids]
    for record in archived:
        patent_id = record["patent_id"]
        match_type = record.get("match_type", "none")
        matched = match_type != "none" and patent_id in verified_ids
        related_form = match_type == "related_form_signature"
        contexts = record.get("contexts") or []
        assignees = "; ".join(record.get("assignees") or []) or "not stated"
        if matched:
            match_note = (
                f"Verified {match_type} in {record.get('section') or 'patent text'}; "
                f"relationship={'related chemical form, not the same material' if related_form else 'same identity text/signature'}; "
                f"signals={','.join(record.get('signals') or []) or 'text_match'}; "
                f"assignee={assignees}; legal status={record.get('legal_status') or 'not stated'}; "
                f"relevant excerpts: {' | '.join(contexts)}"
            )
        else:
            match_note = "Fetched and scanned, but no exact CAS or distinctive PubChem name/synonym text was verified."
        evidence.append(
            EvidenceItem(
                bd=product.bd,
                cas=product.cas,
                source="Google Patents",
                source_type="verified_original_patent" if matched else "patent_candidate",
                title=f"{patent_id}: {record.get('title') or patent_id}",
                url=record["url"],
                snippet=clean_text(
                    f"Priority/publication: {record.get('priority_date') or 'unknown'} / "
                    f"{record.get('publication_date') or 'unknown'}; family={record.get('family_id') or 'unknown'}; "
                    f"assessment: {match_note}",
                    max(800, excerpt_chars + 550),
                ),
                published_date=record.get("publication_date") or (record.get("dates") or [""])[0],
                query=f"Patent full-text verification for PubChem CID {pubchem.get('selected_cid')}",
                identifiers=patent_id,
                relevance_score=int(record.get("score") or 55),
                match_basis=match_type if matched else "full_text_scanned_no_exact_match",
                direct_identifier_match=bool(matched and match_type == "cas_exact"),
                evidence_scope="qualified" if matched and related_form else "direct" if matched else "indirect",
            )
        )
    return {
        "mode": mode,
        "candidate_count": len(patent_ids),
        "deduped_candidate_count": len(deduped_indexed),
        "search_candidate_count": len(dedupe_patent_identifiers(searched_ids, candidate_cap)),
        "searches": search_records,
        "selected_count": len(records),
        "selected_ids": selected_ids,
        "verified_count": len(verified),
        "verified_ids": [record["patent_id"] for record in verified],
        "early_stopped": high_confidence_hits >= 2,
        "cache_hits": cache_hits,
        "network_requests_this_run": network_requests,
        "estimated_incremental_llm_tokens": sum(estimate_tokens(item.snippet) for item in evidence if item.evidence_scope == "direct"),
        "records": records,
    }, evidence


def collect_chembl(product: ProductInput, client: HttpClient) -> tuple[dict[str, Any], list[EvidenceItem]]:
    endpoint = "https://www.ebi.ac.uk/chembl/api/data/molecule.json"
    data, url, _ = client.get_json(
        endpoint,
        params={"molecule_synonyms__molecule_synonym__iexact": product.cas, "limit": 20},
        allow_not_found=True,
    )
    molecules = (data or {}).get("molecules") or []
    # Do not fall back to ChEMBL's fuzzy full-text search for a CAS number.
    # It can return unrelated records with similar digit fragments, which is
    # worse than an explicit "not found" for an evidence-first workflow.
    evidence: list[EvidenceItem] = []
    normalized: list[dict[str, Any]] = []
    for molecule in molecules[:10]:
        chembl_id = clean_text(molecule.get("molecule_chembl_id"))
        if not chembl_id:
            continue
        names = [
            clean_text(item.get("molecule_synonym"))
            for item in (molecule.get("molecule_synonyms") or [])
            if isinstance(item, dict) and clean_text(item.get("molecule_synonym"))
        ]
        exact_cas = any(name.casefold() == product.cas.casefold() for name in names)
        normalized.append(
            {
                "molecule_chembl_id": chembl_id,
                "pref_name": molecule.get("pref_name"),
                "molecule_type": molecule.get("molecule_type"),
                "max_phase": molecule.get("max_phase"),
                "first_approval": molecule.get("first_approval"),
                "therapeutic_flag": molecule.get("therapeutic_flag"),
                "exact_cas_synonym": exact_cas,
                "synonyms": names[:40],
            }
        )
        title = clean_text(molecule.get("pref_name") or chembl_id)
        evidence.append(
            EvidenceItem(
                bd=product.bd,
                cas=product.cas,
                source="ChEMBL",
                source_type="drug_database",
                title=title,
                url=f"https://www.ebi.ac.uk/chembl/explore/compound/{chembl_id}",
                snippet=clean_text(
                    f"{chembl_id}; type={molecule.get('molecule_type')}; max_phase={molecule.get('max_phase')}; "
                    f"first_approval={molecule.get('first_approval')}; exact CAS synonym={exact_cas}; synonyms={'; '.join(names[:8])}",
                    1200,
                ),
                identifiers=chembl_id,
                relevance_score=94 if exact_cas else 72,
                match_basis="cas_exact_synonym" if exact_cas else "structured_cas_query",
                direct_identifier_match=exact_cas,
            )
        )
    return {"query_url": url, "found": bool(normalized), "molecules": normalized}, evidence


def literature_url(record: dict[str, Any]) -> str:
    if record.get("doi"):
        return f"https://doi.org/{record['doi']}"
    if record.get("pmcid"):
        return f"https://pmc.ncbi.nlm.nih.gov/articles/{record['pmcid']}/"
    if record.get("pmid"):
        return f"https://pubmed.ncbi.nlm.nih.gov/{record['pmid']}/"
    source = clean_text(record.get("source"))
    identifier = clean_text(record.get("id"))
    return f"https://europepmc.org/article/{urllib.parse.quote(source)}/{urllib.parse.quote(identifier)}"


def collect_europepmc(
    product: ProductInput, client: HttpClient, fallback_name: str
) -> tuple[dict[str, Any], list[EvidenceItem]]:
    endpoint = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
    queries = [(f'"{product.cas}"', "cas")]
    results: list[dict[str, Any]] = []
    query_urls: list[str] = []
    for query, query_kind in queries:
        data, url, _ = client.get_json(
            endpoint,
            params={"query": query, "format": "json", "resultType": "core", "pageSize": 10},
            allow_not_found=True,
        )
        query_urls.append(url)
        batch = (((data or {}).get("resultList") or {}).get("result") or [])
        for item in batch:
            item["_query"] = query
            item["_query_kind"] = query_kind
        results.extend(batch)
    if not results and fallback_name and fallback_name.casefold() != product.cas.casefold() and len(fallback_name) >= 6:
        query = f'"{fallback_name}"'
        data, url, _ = client.get_json(
            endpoint,
            params={"query": query, "format": "json", "resultType": "core", "pageSize": 10},
            allow_not_found=True,
        )
        query_urls.append(url)
        results = (((data or {}).get("resultList") or {}).get("result") or [])
        for item in results:
            item["_query"] = query
            item["_query_kind"] = "name_fallback"

    evidence: list[EvidenceItem] = []
    normalized: list[dict[str, Any]] = []
    for item in results[:10]:
        record = {
            "source": item.get("source"),
            "id": item.get("id"),
            "pmid": item.get("pmid"),
            "pmcid": item.get("pmcid"),
            "doi": item.get("doi"),
            "title": clean_text(item.get("title")),
            "author_string": clean_text(item.get("authorString")),
            "journal_title": clean_text(item.get("journalTitle")),
            "pub_year": clean_text(item.get("pubYear")),
            "abstract": clean_text(item.get("abstractText"), 5000),
            "query": item.get("_query"),
            "query_kind": item.get("_query_kind"),
        }
        normalized.append(record)
        exact = product.cas.casefold() in json.dumps(item, ensure_ascii=False).casefold()
        indirect = record["query_kind"] == "name_fallback" or not exact
        evidence.append(
            EvidenceItem(
                bd=product.bd,
                cas=product.cas,
                source="Europe PMC",
                source_type="literature",
                title=record["title"] or f"Europe PMC {record['id']}",
                url=literature_url(record),
                snippet=clean_text(
                    f"{record['author_string']}. {record['journal_title']} ({record['pub_year']}). {record['abstract']}",
                    1400,
                ),
                published_date=record["pub_year"],
                query=clean_text(record["query"]),
                identifiers=";".join(
                    part for part in [f"PMID:{record['pmid']}" if record["pmid"] else "", f"DOI:{record['doi']}" if record["doi"] else ""] if part
                ),
                relevance_score=88 if exact else 45,
                match_basis=(
                    "cas_exact"
                    if exact
                    else "name_fallback"
                    if record["query_kind"] == "name_fallback"
                    else "cas_query_unconfirmed"
                ),
                direct_identifier_match=exact,
                evidence_scope="indirect" if indirect else "direct",
            )
        )
    return {"query_urls": query_urls, "found": bool(normalized), "records": normalized}, evidence


def collect_pubmed(product: ProductInput, client: HttpClient, retmax: int, email: str) -> tuple[dict[str, Any], list[EvidenceItem]]:
    common = {"tool": "inquiry_evidence_collector", "email": email}
    if client.ncbi_api_key:
        common["api_key"] = client.ncbi_api_key
    query = f'"{product.cas}"[All Fields]'
    search_params = {"db": "pubmed", "term": query, "retmode": "json", "retmax": retmax, **common}
    search_data, search_url, _ = client.get_json(
        "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi", params=search_params
    )
    ids = (((search_data or {}).get("esearchresult") or {}).get("idlist") or [])
    if not ids:
        return {"query": query, "search_url": search_url, "found": False, "records": []}, []
    fetch_params = {"db": "pubmed", "id": ",".join(ids), "retmode": "xml", **common}
    xml_body, fetch_url, _ = client.get_text(
        "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi", params=fetch_params
    )
    records: list[dict[str, Any]] = []
    evidence: list[EvidenceItem] = []
    root = ET.fromstring(xml_body or "<PubmedArticleSet/>")
    for article in root.findall(".//PubmedArticle"):
        pmid = xml_text(article.find(".//MedlineCitation/PMID"))
        title = xml_text(article.find(".//Article/ArticleTitle"))
        abstract = clean_text(" ".join(xml_text(node) for node in article.findall(".//Article/Abstract/AbstractText")), 5000)
        journal = xml_text(article.find(".//Article/Journal/Title"))
        date_node = article.find(".//Article/Journal/JournalIssue/PubDate")
        published = xml_text(date_node)
        doi = ""
        pmcid = ""
        for article_id in article.findall(".//PubmedData/ArticleIdList/ArticleId"):
            id_type = article_id.attrib.get("IdType", "")
            if id_type == "doi":
                doi = xml_text(article_id)
            elif id_type == "pmc":
                pmcid = xml_text(article_id)
        record = {
            "pmid": pmid,
            "pmcid": pmcid,
            "doi": doi,
            "title": title,
            "journal": journal,
            "published": published,
            "abstract": abstract,
        }
        records.append(record)
        url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else fetch_url
        direct_cas_match = product.cas.casefold() in f"{title} {abstract}".casefold()
        evidence.append(
            EvidenceItem(
                bd=product.bd,
                cas=product.cas,
                source="PubMed",
                source_type="literature",
                title=title or f"PubMed {pmid}",
                url=url,
                snippet=clean_text(f"{journal}. {published}. {abstract}", 1400),
                published_date=extract_year(published),
                query=query,
                identifiers=";".join(part for part in [f"PMID:{pmid}" if pmid else "", f"DOI:{doi}" if doi else ""] if part),
                relevance_score=90,
                match_basis="cas_exact" if direct_cas_match else "cas_indexed_query",
                direct_identifier_match=direct_cas_match,
            )
        )
    return {
        "query": query,
        "search_url": search_url,
        "fetch_url": fetch_url,
        "found": bool(records),
        "records": records,
    }, evidence


def collect_clinicaltrials(
    product: ProductInput, client: HttpClient, fallback_name: str
) -> tuple[dict[str, Any], list[EvidenceItem]]:
    endpoint = "https://clinicaltrials.gov/api/v2/studies"
    queries = [(f'"{product.cas}"', "cas")]
    responses: list[dict[str, Any]] = []
    query_urls: list[str] = []
    for query, kind in queries:
        data, url, _ = client.get_json(
            endpoint,
            params={"query.term": query, "pageSize": 10, "format": "json", "countTotal": "true"},
            allow_not_found=True,
        )
        query_urls.append(url)
        studies = (data or {}).get("studies") or []
        for study in studies:
            responses.append({"query": query, "query_kind": kind, "study": study})
    punctuation_count = sum(1 for character in fallback_name if character in "()[]{}=,;/\\")
    word_count = len(fallback_name.split())
    concise_name = (
        fallback_name
        and fallback_name.casefold() != product.cas.casefold()
        and 5 <= len(fallback_name) <= 80
        and punctuation_count <= 3
        and word_count <= 10
        and re.fullmatch(r"[A-Za-z][A-Za-z '\-]{3,79}", fallback_name) is not None
    )
    if not responses and concise_name:
        query = fallback_name
        data, url, _ = client.get_json(
            endpoint,
            params={"query.intr": query, "pageSize": 10, "format": "json", "countTotal": "true"},
            allow_not_found=True,
        )
        query_urls.append(url)
        for study in (data or {}).get("studies") or []:
            responses.append({"query": query, "query_kind": "name_fallback", "study": study})

    records: list[dict[str, Any]] = []
    evidence: list[EvidenceItem] = []
    for wrapper in responses[:10]:
        study = wrapper["study"]
        protocol = study.get("protocolSection") or {}
        identification = protocol.get("identificationModule") or {}
        status = protocol.get("statusModule") or {}
        conditions = protocol.get("conditionsModule") or {}
        sponsors = protocol.get("sponsorCollaboratorsModule") or {}
        arms = protocol.get("armsInterventionsModule") or {}
        nct_id = clean_text(identification.get("nctId"))
        record = {
            "nct_id": nct_id,
            "brief_title": clean_text(identification.get("briefTitle")),
            "official_title": clean_text(identification.get("officialTitle")),
            "overall_status": clean_text(status.get("overallStatus")),
            "start_date": clean_text((status.get("startDateStruct") or {}).get("date")),
            "completion_date": clean_text((status.get("completionDateStruct") or {}).get("date")),
            "conditions": conditions.get("conditions") or [],
            "lead_sponsor": clean_text((sponsors.get("leadSponsor") or {}).get("name")),
            "interventions": [
                {
                    "type": clean_text(item.get("type")),
                    "name": clean_text(item.get("name")),
                    "other_names": item.get("otherNames") or [],
                }
                for item in (arms.get("interventions") or [])
            ],
            "query": wrapper["query"],
            "query_kind": wrapper["query_kind"],
        }
        records.append(record)
        exact = product.cas.casefold() in json.dumps(study, ensure_ascii=False).casefold()
        indirect = wrapper["query_kind"] == "name_fallback" or not exact
        intervention_names = "; ".join(item["name"] for item in record["interventions"] if item["name"])
        evidence.append(
            EvidenceItem(
                bd=product.bd,
                cas=product.cas,
                source="ClinicalTrials.gov",
                source_type="clinical_trial",
                title=record["brief_title"] or nct_id,
                url=f"https://clinicaltrials.gov/study/{nct_id}" if nct_id else query_urls[-1],
                snippet=clean_text(
                    f"status={record['overall_status']}; sponsor={record['lead_sponsor']}; conditions={'; '.join(record['conditions'])}; "
                    f"interventions={intervention_names}; start={record['start_date']}; completion={record['completion_date']}",
                    1400,
                ),
                published_date=extract_year(record["start_date"]),
                query=record["query"],
                identifiers=nct_id,
                relevance_score=94 if exact else 45,
                match_basis=(
                    "cas_exact"
                    if exact
                    else "name_fallback"
                    if wrapper["query_kind"] == "name_fallback"
                    else "cas_query_unconfirmed"
                ),
                direct_identifier_match=exact,
                evidence_scope="indirect" if indirect else "direct",
            )
        )
    return {"query_urls": query_urls, "found": bool(records), "records": records}, evidence


def web_domain_score(url: str) -> int:
    host = (urllib.parse.urlsplit(url).hostname or "").lower()
    high = (
        ".gov",
        "who.int",
        "ema.europa.eu",
        "wipo.int",
        "epo.org",
        "nature.com",
        "science.org",
        "pubmed.ncbi.nlm.nih.gov",
        "pmc.ncbi.nlm.nih.gov",
        "clinicaltrials.gov",
        "patents.google.com",
    )
    medium = ("doi.org", "springer.com", "sciencedirect.com", "acs.org", "rsc.org", "company")
    if any(token in host for token in high):
        return 85
    if any(token in host for token in medium):
        return 72
    return 52


def quoted_search_term(value: str) -> str:
    """Return a bounded phrase safe to embed in a search query."""
    return clean_text(value, 240).replace('"', " ").strip()


def distinctive_primary_name_match(url: str, name: str) -> bool:
    """Allow name-only direct evidence only for distinctive names on primary domains."""
    normalized = clean_text(name)
    tokens = re.findall(r"[A-Za-z0-9]+", normalized)
    if len(normalized) < 18 or len(tokens) < 2:
        return False
    host = (urlparse(url).hostname or "").casefold()
    primary_hosts = (
        "patents.google.com",
        "wipo.int",
        "epo.org",
        "pubmed.ncbi.nlm.nih.gov",
        "pmc.ncbi.nlm.nih.gov",
        "doi.org",
    )
    return any(host == item or host.endswith("." + item) for item in primary_hosts)


def collect_tavily(
    product: ProductInput,
    client: HttpClient,
    fallback_name: str,
    api_key: str | None,
    count: int,
    include_recent_events: bool = False,
) -> tuple[dict[str, Any], list[EvidenceItem]]:
    if not api_key:
        return {"skipped": True, "reason": "TAVILY_API_KEY not set", "queries": []}, []
    current_year = dt.date.today().year
    english_name = quoted_search_term(fallback_name)
    identity_query = f'("{product.cas}" OR "{english_name}")' if english_name else f'"{product.cas}"'
    queries: list[tuple[str, str, bool]] = [
        (f'"{product.cas}"', "exact_cas", True),
        (
            f'{identity_query} (patent OR intermediate OR impurity OR metabolite OR synthesis OR standard)',
            "use_and_patent",
            True,
        ),
    ]
    if include_recent_events and fallback_name and fallback_name.casefold() != product.cas.casefold():
        queries.append(
            (
                f'"{fallback_name}" (clinical trial OR approval OR tender OR acquisition OR regulatory OR patent) '
                f"({current_year - 1} OR {current_year})",
                "recent_event",
                True,
            )
        )

    endpoint = "https://api.tavily.com/search"
    headers = {"Authorization": f"Bearer {api_key}"}
    records: list[dict[str, Any]] = []
    evidence: list[EvidenceItem] = []
    response_credits = 0.0
    credits_this_run = 0.0
    cache_hits = 0
    for query, query_kind, exact_match in queries:
        payload = {
            "query": query,
            "search_depth": "basic",
            "max_results": min(20, max(1, count)),
            "topic": "general",
            "include_answer": False,
            "include_raw_content": False,
            "include_images": False,
            "auto_parameters": False,
            "exact_match": exact_match,
            "include_usage": True,
        }
        data, url, cached = client.post_json(endpoint, payload=payload, headers=headers)
        usage_credits = float(((data or {}).get("usage") or {}).get("credits") or 0)
        response_credits += usage_credits
        if cached:
            cache_hits += 1
        else:
            credits_this_run += usage_credits
        for result in (data or {}).get("results") or []:
            result_url = clean_text(result.get("url"))
            title = clean_text(result.get("title"))
            content = clean_text(result.get("content"))
            published_date = clean_text(result.get("published_date"))
            try:
                provider_score = float(result.get("score") or 0)
            except (TypeError, ValueError):
                provider_score = 0.0
            record = {
                "query": query,
                "query_kind": query_kind,
                "query_url": url,
                "cached": cached,
                "title": title,
                "url": result_url,
                "content": content,
                "published_date": published_date,
                "provider_score": provider_score,
            }
            records.append(record)
            result_text = f"{title} {content}".casefold()
            exact = product.cas.casefold() in result_text
            exact_name = bool(fallback_name and fallback_name.casefold() in result_text)
            primary_name_match = bool(
                query_kind != "recent_event"
                and exact_name
                and distinctive_primary_name_match(result_url, fallback_name)
            )
            if exact and query_kind != "recent_event":
                score = web_domain_score(result_url) + min(10, max(0, round(provider_score * 10))) + 10
                match_basis = "cas_exact"
                evidence_scope = "direct"
                source_type = "web_search"
            elif primary_name_match:
                score = min(
                    90,
                    web_domain_score(result_url)
                    + min(10, max(0, round(provider_score * 10))),
                )
                match_basis = "name_exact_primary"
                evidence_scope = "direct"
                source_type = "web_search_name_exact"
            else:
                # A reputable domain cannot compensate for a missing product
                # identifier. Keep name-only results in the audit archive but
                # exclude them from the default LLM evidence pack.
                score = min(45, 30 + min(10, max(0, round(provider_score * 10))))
                match_basis = "name_exact" if exact_name else "semantic_only"
                evidence_scope = "indirect"
                source_type = "web_search_indirect"
            evidence.append(
                EvidenceItem(
                    bd=product.bd,
                    cas=product.cas,
                    source="Tavily Search",
                    source_type=source_type,
                    title=title or result_url,
                    url=result_url,
                    snippet=clean_text(content, 1400),
                    published_date=published_date,
                    query=query,
                    relevance_score=min(99, score),
                    match_basis=match_basis,
                    direct_identifier_match=exact,
                    evidence_scope=evidence_scope,
                )
            )
    return {
        "skipped": False,
        "queries": [{"query": query, "query_kind": kind, "exact_match": exact} for query, kind, exact in queries],
        "query_count": len(queries),
        "recent_events_enabled": include_recent_events,
        "cache_hits": cache_hits,
        "usage_credits_in_responses": response_credits,
        "usage_credits_this_run": credits_this_run,
        "records": records,
    }, evidence


def collect_mce(
    product: ProductInput,
    client: HttpClient,
    fallback_name: str,
    api_key: str | None,
    count: int,
) -> tuple[dict[str, Any], list[EvidenceItem]]:
    """Collect CAS-bound, MCE-curated pharmacology profiles via Tavily domain search."""
    if not api_key:
        return {
            "skipped": True,
            "reason": "TAVILY_API_KEY not set; MCE domain search uses Tavily",
            "queries": [],
        }, []

    english_name = quoted_search_term(fallback_name)
    identity_query = f'("{product.cas}" OR "{english_name}")' if english_name else f'"{product.cas}"'
    query = (
        f'{identity_query} (target OR inhibitor OR agonist OR mechanism OR pathway OR "biological activity" OR IC50)'
    )
    payload = {
        "query": query,
        "search_depth": "basic",
        "max_results": min(5, max(1, count)),
        "topic": "general",
        "include_answer": False,
        "include_raw_content": False,
        "include_images": False,
        "include_domains": ["medchemexpress.com"],
        "auto_parameters": False,
        "exact_match": False,
        "include_usage": True,
    }
    data, query_url, cached = client.post_json(
        "https://api.tavily.com/search",
        payload=payload,
        headers={"Authorization": f"Bearer {api_key}"},
    )
    usage_credits = float(((data or {}).get("usage") or {}).get("credits") or 0)
    records: list[dict[str, Any]] = []
    evidence: list[EvidenceItem] = []
    for result in (data or {}).get("results") or []:
        result_url = clean_text(result.get("url"))
        host = (urlparse(result_url).hostname or "").casefold()
        if not (host == "medchemexpress.com" or host.endswith(".medchemexpress.com")):
            continue
        title = clean_text(result.get("title"))
        content = clean_text(result.get("content"))
        published_date = clean_text(result.get("published_date"))
        try:
            provider_score = float(result.get("score") or 0)
        except (TypeError, ValueError):
            provider_score = 0.0
        result_text = f"{title} {content}".casefold()
        exact_cas = product.cas.casefold() in result_text
        exact_name = bool(fallback_name and fallback_name.casefold() in result_text)
        if exact_cas:
            relevance_score = min(94, 82 + min(10, max(0, round(provider_score * 10))))
            match_basis = "cas_exact_mce"
            evidence_scope = "direct"
            source_type = "mce_product_profile"
        else:
            relevance_score = min(48, 34 + min(10, max(0, round(provider_score * 10))))
            match_basis = "name_exact_mce" if exact_name else "semantic_only_mce"
            evidence_scope = "indirect"
            source_type = "mce_product_profile_indirect"
        record = {
            "query": query,
            "query_url": query_url,
            "cached": cached,
            "title": title,
            "url": result_url,
            "content": content,
            "published_date": published_date,
            "provider_score": provider_score,
            "exact_cas": exact_cas,
            "exact_name": exact_name,
        }
        records.append(record)
        evidence.append(
            EvidenceItem(
                bd=product.bd,
                cas=product.cas,
                source="MedChemExpress (MCE)",
                source_type=source_type,
                title=title or result_url,
                url=result_url,
                snippet=clean_text(content, 1800),
                published_date=published_date,
                query=query,
                identifiers=product.cas if exact_cas else "",
                relevance_score=relevance_score,
                match_basis=match_basis,
                direct_identifier_match=exact_cas,
                evidence_scope=evidence_scope,
            )
        )
    return {
        "skipped": False,
        "queries": [
            {
                "query": query,
                "include_domains": ["medchemexpress.com"],
                "exact_match": False,
            }
        ],
        "query_count": 1,
        "cache_hits": 1 if cached else 0,
        "usage_credits_in_responses": usage_credits,
        "usage_credits_this_run": 0.0 if cached else usage_credits,
        "evidence_policy": "curated_secondary; exact CAS required for default evidence pack",
        "records": records,
    }, evidence


def collect_brave(
    product: ProductInput,
    client: HttpClient,
    fallback_name: str,
    api_key: str | None,
    count: int,
    include_recent_events: bool = False,
) -> tuple[dict[str, Any], list[EvidenceItem]]:
    if not api_key:
        return {"skipped": True, "reason": "BRAVE_SEARCH_API_KEY not set", "queries": []}, []
    current_year = dt.date.today().year
    english_name = quoted_search_term(fallback_name)
    identity_query = f'("{product.cas}" OR "{english_name}")' if english_name else f'"{product.cas}"'
    queries: list[tuple[str, str]] = [
        (f'"{product.cas}"', "exact_cas"),
        (f'{identity_query} (patent OR intermediate OR impurity OR metabolite OR synthesis OR standard)', "use_and_patent"),
    ]
    if include_recent_events and fallback_name and fallback_name.casefold() != product.cas.casefold():
        queries.append(
            (
                f'"{fallback_name}" (clinical trial OR approval OR tender OR acquisition OR regulatory OR patent) '
                f"({current_year - 1} OR {current_year})",
                "recent_event",
            )
        )
    records: list[dict[str, Any]] = []
    evidence: list[EvidenceItem] = []
    headers = {"Accept": "application/json", "X-Subscription-Token": api_key}
    for query, query_kind in queries:
        data, url, _ = client.get_json(
            "https://api.search.brave.com/res/v1/web/search",
            params={
                "q": query,
                "count": min(20, max(1, count)),
                "search_lang": "en",
                "safesearch": "moderate",
                "spellcheck": "false",
                "extra_snippets": "true",
                "result_filter": "web",
            },
            headers=headers,
        )
        for result in (((data or {}).get("web") or {}).get("results") or []):
            result_url = clean_text(result.get("url"))
            title = clean_text(result.get("title"))
            description = clean_text(result.get("description"))
            extra = [clean_text(item) for item in (result.get("extra_snippets") or []) if clean_text(item)]
            age = clean_text(result.get("age") or result.get("page_age"))
            record = {
                "query": query,
                "query_kind": query_kind,
                "query_url": url,
                "title": title,
                "url": result_url,
                "description": description,
                "extra_snippets": extra,
                "age": age,
            }
            records.append(record)
            result_text = f"{title} {description} {' '.join(extra)}".casefold()
            exact = product.cas.casefold() in result_text
            exact_name = bool(fallback_name and fallback_name.casefold() in result_text)
            primary_name_match = bool(
                query_kind != "recent_event"
                and exact_name
                and distinctive_primary_name_match(result_url, fallback_name)
            )
            if exact and query_kind != "recent_event":
                score = web_domain_score(result_url) + 10
                match_basis = "cas_exact"
                evidence_scope = "direct"
                source_type = "web_search"
            elif primary_name_match:
                score = min(90, web_domain_score(result_url) + 5)
                match_basis = "name_exact_primary"
                evidence_scope = "direct"
                source_type = "web_search_name_exact"
            else:
                score = 38
                match_basis = "name_exact" if exact_name else "semantic_only"
                evidence_scope = "indirect"
                source_type = "web_search_indirect"
            evidence.append(
                EvidenceItem(
                    bd=product.bd,
                    cas=product.cas,
                    source="Brave Search",
                    source_type=source_type,
                    title=title or result_url,
                    url=result_url,
                    snippet=clean_text(" | ".join([description, *extra]), 1400),
                    published_date=age,
                    query=query,
                    relevance_score=min(99, score),
                    match_basis=match_basis,
                    direct_identifier_match=exact,
                    evidence_scope=evidence_scope,
                )
            )
    return {
        "skipped": False,
        "queries": queries,
        "recent_events_enabled": include_recent_events,
        "records": records,
    }, evidence


def canonical_url(url: str) -> str:
    if not url:
        return ""
    parsed = urllib.parse.urlsplit(url)
    query = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
    tracking = {"utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "gclid", "fbclid"}
    filtered = [(key, value) for key, value in query if key.casefold() not in tracking]
    return urllib.parse.urlunsplit((parsed.scheme.lower(), parsed.netloc.lower(), parsed.path.rstrip("/"), urllib.parse.urlencode(filtered), ""))


def rank_and_dedupe(items: Iterable[EvidenceItem], cas: str) -> list[EvidenceItem]:
    current_year = dt.date.today().year
    best: dict[str, EvidenceItem] = {}
    for item in items:
        if not item.url:
            continue
        haystack = f"{item.title} {item.snippet} {item.identifiers}".casefold()
        if cas.casefold() in haystack:
            item.relevance_score = min(100, item.relevance_score + 6)
        year = extract_year(item.published_date)
        if year and int(year) >= current_year - 1:
            item.relevance_score = min(100, item.relevance_score + 3)
        doi_match = re.search(r"DOI:([^;\s]+)", item.identifiers, re.IGNORECASE)
        pmid_match = re.search(r"PMID:(\d+)", item.identifiers, re.IGNORECASE)
        if doi_match:
            key = f"doi:{doi_match.group(1).casefold().rstrip('.')}"
        elif pmid_match:
            key = f"pmid:{pmid_match.group(1)}"
        else:
            key = canonical_url(item.url) or f"{item.source}|{item.title}".casefold()
        existing = best.get(key)
        if existing is None or item.relevance_score > existing.relevance_score:
            best[key] = item
    return sorted(best.values(), key=lambda item: (-item.relevance_score, item.source, item.title.casefold()))


def select_pack_items(items: list[EvidenceItem], limit: int) -> list[EvidenceItem]:
    if limit <= 0:
        return []
    # Keep indirect name fallbacks and unconfirmed patent/index hits in the
    # full CSV/raw archive, but do not feed them to the summarizing model by
    # default. This separation preserves recall without sacrificing precision.
    eligible = [item for item in items if item.evidence_scope != "indirect"]
    if not eligible:
        return []

    source_caps = {
        "Tavily Search": 4,
        "Brave Search": 4,
        "MedChemExpress (MCE)": 2,
        "Google Patents": 5,
        "Europe PMC": 3,
        "PubMed": 3,
        "ClinicalTrials.gov": 2,
        "PubChem patent cross-references": 1,
    }
    type_caps = {
        "chemical_identity": 1,
        "web_search": 4,
        "original_patent": 4,
        "verified_original_patent": 5,
        "patent_index": 1,
        "literature": 4,
        "clinical_trial": 2,
        "drug_database": 2,
        "drug_annotation": 2,
        "mce_product_profile": 2,
    }
    selected: list[EvidenceItem] = []
    used: set[int] = set()
    per_type: defaultdict[str, int] = defaultdict(int)
    per_source: defaultdict[str, int] = defaultdict(int)
    web_hosts: set[str] = set()

    def can_select(item: EvidenceItem) -> bool:
        if per_source[item.source] >= source_caps.get(item.source, 4):
            return False
        if per_type[item.source_type] >= type_caps.get(item.source_type, 4):
            return False
        if item.source_type == "web_search":
            host = (urllib.parse.urlsplit(item.url).hostname or "").casefold()
            if host and host in web_hosts:
                return False
        return True

    def add_item(index: int, item: EvidenceItem) -> None:
        selected.append(item)
        used.add(index)
        per_type[item.source_type] += 1
        per_source[item.source] += 1
        if item.source_type == "web_search":
            host = (urllib.parse.urlsplit(item.url).hostname or "").casefold()
            if host:
                web_hosts.add(host)

    # First preserve source diversity. Caps apply even when the total number
    # of candidates is below the requested pack limit.
    for idx, item in enumerate(eligible):
        if per_source[item.source] == 0 and can_select(item):
            add_item(idx, item)
            if len(selected) >= limit:
                return sorted(selected, key=lambda value: (-value.relevance_score, value.source, value.title.casefold()))

    # Then fill by score while enforcing per-source, per-type and per-domain
    # limits so one search provider or supplier site cannot dominate.
    for idx, item in enumerate(eligible):
        if idx in used or not can_select(item):
            continue
        add_item(idx, item)
        if len(selected) >= limit:
            break
    return sorted(selected, key=lambda value: (-value.relevance_score, value.source, value.title.casefold()))


def safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._") or "unknown"


def compact_product_record(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "input": record["input"],
        "cas_valid": record["cas_valid"],
        "identity": record["identity"],
        "source_status": record["source_status"],
        "evidence_count": len(record["evidence_items"]),
        "raw_file": record["raw_file"],
    }


def identity_from_pubchem(pubchem: dict[str, Any]) -> dict[str, Any]:
    properties = pubchem.get("properties") or {}
    synonyms = pubchem.get("synonyms") or []
    return {
        "pubchem_cid": pubchem.get("selected_cid"),
        "title": clean_text(properties.get("Title") or properties.get("IUPACName")),
        "iupac_name": clean_text(properties.get("IUPACName")),
        "molecular_formula": clean_text(properties.get("MolecularFormula")),
        "molecular_weight": properties.get("MolecularWeight"),
        "smiles": clean_text(properties.get("SMILES") or properties.get("ConnectivitySMILES")),
        "inchi": clean_text(properties.get("InChI")),
        "inchikey": clean_text(properties.get("InChIKey")),
        "exact_mass": properties.get("ExactMass"),
        "synonyms": synonyms[:25],
    }


def collect_one(
    product: ProductInput,
    client: HttpClient,
    sources: set[str],
    args: argparse.Namespace,
) -> dict[str, Any]:
    source_data: dict[str, Any] = {}
    source_status: dict[str, Any] = {}
    all_evidence: list[EvidenceItem] = []

    def run_source(name: str, function: Any, *function_args: Any) -> Any:
        if name not in sources:
            source_status[name] = {"status": "disabled", "items": 0}
            return {}
        try:
            data, evidence = function(*function_args)
            source_data[name] = data
            all_evidence.extend(evidence)
            skipped = bool(data.get("skipped")) if isinstance(data, dict) else False
            source_status[name] = {
                "status": "skipped" if skipped else "ok",
                "items": len(evidence),
                "reason": data.get("reason", "") if isinstance(data, dict) else "",
            }
            if isinstance(data, dict) and "usage_credits_this_run" in data:
                source_status[name]["usage_credits_this_run"] = data["usage_credits_this_run"]
                source_status[name]["cache_hits"] = data.get("cache_hits", 0)
            if isinstance(data, dict) and "network_requests_this_run" in data:
                source_status[name]["network_requests_this_run"] = data["network_requests_this_run"]
                source_status[name]["cache_hits"] = data.get("cache_hits", 0)
            return data
        except Exception as exc:  # Continue other sources; preserve the exact source error.
            source_data[name] = {"error": f"{type(exc).__name__}: {exc}"}
            source_status[name] = {"status": "error", "items": 0, "error": f"{type(exc).__name__}: {exc}"}
            return {}

    pubchem = run_source("pubchem", collect_pubchem, product, client)
    identity = identity_from_pubchem(pubchem) if pubchem else {
        "pubchem_cid": None,
        "title": "",
        "iupac_name": "",
        "molecular_formula": "",
        "molecular_weight": None,
        "smiles": "",
        "inchi": "",
        "inchikey": "",
        "exact_mass": None,
        "synonyms": [],
    }
    fallback_name = identity.get("title") or ""
    run_source(
        "googlepatents",
        collect_google_patents,
        product,
        client,
        pubchem,
        fallback_name,
        getattr(args, "patent_verification", "balanced"),
        getattr(args, "patent_pages", 0),
        getattr(args, "patent_candidate_cap", 300),
        getattr(args, "patent_snippet_chars", 1200),
    )
    run_source("chembl", collect_chembl, product, client)
    run_source("europepmc", collect_europepmc, product, client, fallback_name)
    run_source("pubmed", collect_pubmed, product, client, args.literature_limit, args.ncbi_email)
    run_source("clinicaltrials", collect_clinicaltrials, product, client, fallback_name)
    run_source(
        "tavily",
        collect_tavily,
        product,
        client,
        fallback_name,
        args.tavily_api_key,
        args.web_results,
        args.include_recent_events,
    )
    if "mce" in sources and product.row_number not in args.mce_allowed_rows:
        source_data["mce"] = {
            "skipped": True,
            "reason": f"MCE optional cap: only first {args.mce_max_products} products are queried per run",
        }
        source_status["mce"] = {
            "status": "skipped",
            "items": 0,
            "reason": source_data["mce"]["reason"],
        }
    else:
        run_source(
            "mce",
            collect_mce,
            product,
            client,
            fallback_name,
            args.tavily_api_key,
            args.web_results,
        )
    run_source(
        "brave",
        collect_brave,
        product,
        client,
        fallback_name,
        args.brave_api_key,
        args.web_results,
        args.include_recent_events,
    )

    ranked = rank_and_dedupe(all_evidence, product.cas)
    return {
        "input": dataclasses.asdict(product),
        "cas_valid": cas_is_valid(product.cas),
        "identity": identity,
        "source_status": source_status,
        "source_data": source_data,
        "evidence_items": [dataclasses.asdict(item) for item in ranked],
    }


def evidence_item_from_dict(value: dict[str, Any]) -> EvidenceItem:
    return EvidenceItem(**{field.name: value.get(field.name, field.default) for field in dataclasses.fields(EvidenceItem)})


def build_markdown_pack(
    records: list[dict[str, Any]],
    pack_items: int,
    snippet_chars: int,
    patent_snippet_chars: int = 1200,
) -> tuple[str, list[dict[str, Any]]]:
    lines = [
        "# 询单产品证据包（脚本采集，尚未由大模型下结论）",
        "",
        f"生成时间：{utc_now_iso()}",
        "",
        "使用要求：逐项核验身份与来源；区分直接证据、商业目录和结构推断；PubChem 专利索引仅是候选，只有标为 verified_original_patent 且给出正文命中范围/片段的条目才属于正文二次核验证据；不得把相关化学形式自动当作同一物料。",
        "",
    ]
    manifest: list[dict[str, Any]] = []
    for index, record in enumerate(records, start=1):
        product = record["input"]
        identity = record["identity"]
        items = [evidence_item_from_dict(item) for item in record["evidence_items"]]
        selected = select_pack_items(items, pack_items)
        indirect_count = sum(1 for item in items if item.evidence_scope == "indirect")
        start = len("\n".join(lines))
        lines.extend(
            [
                f"## {index}. {product.get('bd') or '无BD'} | CAS {product['cas']}",
                "",
                f"- CAS 校验：{'通过' if record['cas_valid'] else '未通过'}",
                f"- PubChem CID：{identity.get('pubchem_cid') or '未找到'}",
                f"- 首选名称：{identity.get('title') or '未找到'}",
                f"- IUPAC 名称：{identity.get('iupac_name') or '未找到'}",
                f"- 分子式 / 分子量：{identity.get('molecular_formula') or '未找到'} / {identity.get('molecular_weight') or '未找到'}",
                f"- InChIKey：{identity.get('inchikey') or '未找到'}",
                f"- 同义词（截取）：{'; '.join(identity.get('synonyms') or []) or '未找到'}",
                "",
                "### 候选证据",
                "",
            ]
        )
        if not selected:
            lines.extend(["未采集到候选证据。请检查各数据源状态和网络连接。", ""])
        for evidence_index, item in enumerate(selected, start=1):
            lines.extend(
                [
                    f"#### E{evidence_index:02d} | {item.source} | {item.source_type} | 相关性分数 {item.relevance_score}",
                    "",
                    f"- 标题：{item.title}",
                    f"- 日期：{item.published_date or '未标注'}",
                    f"- 标识符：{item.identifiers or '无'}",
                    f"- 匹配依据：{item.match_basis or '未标注'}",
                    f"- 证据范围：{item.evidence_scope or '未标注'}",
                    f"- 直接 CAS 匹配：{'是' if item.direct_identifier_match else '否（结构化来源或名称/同义词直接证据）'}",
                    f"- 检索式：{item.query or '结构化接口直接查询'}",
                    f"- URL：{item.url}",
                    f"- 摘要：{clean_text(item.snippet, patent_snippet_chars if item.source_type == 'verified_original_patent' else snippet_chars) or '无摘要'}",
                    "",
                ]
            )
        source_states = "; ".join(
            f"{name}={status.get('status')}({status.get('items', 0)})" for name, status in record["source_status"].items()
        )
        lines.extend([f"数据源状态：{source_states}", "", "---", ""])
        end = len("\n".join(lines))
        section_text = "\n".join(lines)[start:end]
        manifest.append(
            {
                "bd": product.get("bd", ""),
                "cas": product["cas"],
                "selected_evidence_items": len(selected),
                "direct_candidate_items": len(items) - indirect_count,
                "excluded_indirect_items": indirect_count,
                "section_characters": len(section_text),
                "estimated_tokens": estimate_tokens(section_text),
            }
        )
    text = "\n".join(lines).rstrip() + "\n"
    return text, manifest


def write_outputs(records: list[dict[str, Any]], output_dir: Path, args: argparse.Namespace, started_at: str) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    raw_dir = output_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    for record in records:
        product = record["input"]
        raw_path = raw_dir / f"{safe_filename(product.get('bd') or 'no-bd')}__{safe_filename(product['cas'])}.json"
        raw_path.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        record["raw_file"] = str(raw_path.relative_to(output_dir)).replace("\\", "/")

    with (output_dir / "products.jsonl").open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            handle.write(json.dumps(compact_product_record(record), ensure_ascii=False) + "\n")

    evidence_fields = [field.name for field in dataclasses.fields(EvidenceItem)]
    with (output_dir / "evidence_candidates.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=evidence_fields)
        writer.writeheader()
        for record in records:
            writer.writerows(record["evidence_items"])

    pack_text, token_manifest = build_markdown_pack(
        records,
        args.pack_items,
        args.snippet_chars,
        args.patent_snippet_chars,
    )
    candidate_total = sum(len(record["evidence_items"]) for record in records)
    indirect_candidate_total = sum(
        1
        for record in records
        for item in record["evidence_items"]
        if item.get("evidence_scope") == "indirect"
    )
    selected_total = sum(int(item["selected_evidence_items"]) for item in token_manifest)
    (output_dir / "llm_evidence_pack.md").write_text(pack_text, encoding="utf-8")
    (output_dir / "llm_token_estimates.json").write_text(
        json.dumps(
            {
                "method": "CJK characters + ceil(other characters / 4); estimate only, not account billing data",
                "total_characters": len(pack_text),
                "estimated_total_tokens": estimate_tokens(pack_text),
                "candidate_evidence_items": candidate_total,
                "excluded_indirect_items": indirect_candidate_total,
                "selected_evidence_items": selected_total,
                "products": token_manifest,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    error_count = sum(
        1 for record in records for status in record["source_status"].values() if status.get("status") == "error"
    )
    metadata = {
        "script_version": SCRIPT_VERSION,
        "started_at": started_at,
        "finished_at": utc_now_iso(),
        "input_mode": "single_cas" if args.cas else "file",
        "input_file": str(args.input.resolve()) if args.input else "",
        "input_cas": args.cas or "",
        "product_count": len(records),
        "sources": sorted(args.sources),
        "source_error_count": error_count,
        "candidate_evidence_items": candidate_total,
        "indirect_candidate_items": indirect_candidate_total,
        "selected_pack_items": selected_total,
        "include_recent_events": args.include_recent_events,
        "patent_verification_mode": args.patent_verification,
        "patent_candidate_cap": args.patent_candidate_cap,
        "patent_page_limit_override": args.patent_pages,
        "patent_verified_items": sum(
            int(((record.get("source_data") or {}).get("googlepatents") or {}).get("verified_count") or 0)
            for record in records
        ),
        "patent_network_requests_this_run": sum(
            int(((record.get("source_data") or {}).get("googlepatents") or {}).get("network_requests_this_run") or 0)
            for record in records
        ),
        "patent_cache_hits": sum(
            int(((record.get("source_data") or {}).get("googlepatents") or {}).get("cache_hits") or 0)
            for record in records
        ),
        "patent_estimated_incremental_llm_tokens": sum(
            int(((record.get("source_data") or {}).get("googlepatents") or {}).get("estimated_incremental_llm_tokens") or 0)
            for record in records
        ),
        "tavily_enabled": bool(args.tavily_api_key and "tavily" in args.sources),
        "tavily_credits_this_run": sum(
            sum(
                float(((record.get("source_data") or {}).get(source_name) or {}).get("usage_credits_this_run") or 0)
                for source_name in ("tavily", "mce")
            )
            for record in records
        ),
        "tavily_cache_hits": sum(
            sum(
                int(((record.get("source_data") or {}).get(source_name) or {}).get("cache_hits") or 0)
                for source_name in ("tavily", "mce")
            )
            for record in records
        ),
        "mce_enabled": bool(args.tavily_api_key and "mce" in args.sources),
        "mce_product_cap": args.mce_max_products,
        "mce_products_eligible": (
            min(len(records), args.mce_max_products)
            if "mce" in args.sources and args.mce_max_products > 0
            else (len(records) if "mce" in args.sources else 0)
        ),
        "mce_products_skipped_by_optional_cap": (
            max(0, len(records) - args.mce_max_products)
            if "mce" in args.sources and args.mce_max_products > 0
            else 0
        ),
        "mce_tavily_credits_this_run": sum(
            float(((record.get("source_data") or {}).get("mce") or {}).get("usage_credits_this_run") or 0)
            for record in records
        ),
        "mce_tavily_cache_hits": sum(
            int(((record.get("source_data") or {}).get("mce") or {}).get("cache_hits") or 0)
            for record in records
        ),
        "brave_enabled": bool(args.brave_api_key and "brave" in args.sources),
        "offline": args.offline,
        "cache_dir": str(args.cache_dir.resolve()),
        "resumed_from_checkpoint": bool(getattr(args, "resumed_from_checkpoint", False)),
        "checkpoint_products_reused": int(getattr(args, "checkpoint_products_reused", 0)),
        "resume_count": int(getattr(args, "resume_count", 0)),
        "resume_state": RESUME_STATE_FILENAME,
        "outputs": {
            "compact_products": "products.jsonl",
            "evidence_candidates": "evidence_candidates.csv",
            "llm_pack": "llm_evidence_pack.md",
            "token_estimates": "llm_token_estimates.json",
            "raw_records": "raw/",
        },
    }
    (output_dir / "run_metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="从询单 Excel/CSV 批量收集化学品身份、文献、专利、临床和网页候选证据；不调用大模型。"
    )
    parser.add_argument(
        "input",
        type=Path,
        nargs="?",
        help="包含 BD/CAS 列的 .xlsx、.xlsm 或 .csv 文件；与 --cas 二选一",
    )
    parser.add_argument("--cas", default="", help="直接处理单个 CAS，无需 Excel/CSV")
    parser.add_argument(
        "--resume-dir",
        type=Path,
        default=None,
        help="从未完成输出目录的产品检查点续跑；不要再提供输入文件或 --cas",
    )
    parser.add_argument("--output-dir", type=Path, default=None, help="输出目录；默认 outputs/evidence-<时间>")
    parser.add_argument("--sheet", default=None, help="Excel 工作表名称；默认第一个工作表")
    parser.add_argument("--bd-column", default="BD", help="BD 列标题，默认 BD")
    parser.add_argument("--cas-column", default="CAS", help="CAS 列标题，默认 CAS")
    parser.add_argument("--sources", default=",".join(DEFAULT_SOURCES), help=f"启用的数据源，逗号分隔；可选：{','.join(SUPPORTED_SOURCES)}")
    parser.add_argument("--limit", type=int, default=0, help="仅处理前 N 个产品，用于测试；0 表示全部")
    parser.add_argument("--max-workers", type=int, default=4, help="产品并发数，默认 4")
    parser.add_argument("--timeout", type=float, default=35.0, help="单次 HTTP 请求超时秒数，默认 35")
    parser.add_argument("--retries", type=int, default=3, help="可重试错误的重试次数，默认 3")
    parser.add_argument("--cache-dir", type=Path, default=Path(".cache/inquiry_evidence"), help="HTTP 缓存目录")
    parser.add_argument("--cache-ttl-days", type=float, default=14.0, help="缓存有效天数，默认 14")
    parser.add_argument("--offline", action="store_true", help="只使用已有缓存，不联网")
    parser.add_argument("--literature-limit", type=int, default=8, help="PubMed 最多返回多少篇，默认 8")
    parser.add_argument(
        "--patent-verification",
        choices=tuple(PATENT_VERIFICATION_MODES),
        default="balanced",
        help="专利正文二次核验：off/balanced/deep；默认 balanced",
    )
    parser.add_argument(
        "--patent-pages",
        type=int,
        default=0,
        help="覆盖核验模式的每产品专利正文页数；0 表示使用模式默认值",
    )
    parser.add_argument(
        "--patent-candidate-cap",
        type=int,
        default=300,
        help="每产品进入轻量去重/排序的原始专利号上限，默认 300",
    )
    parser.add_argument(
        "--patent-snippet-chars",
        type=int,
        default=1200,
        help="每篇入选核验专利写入大模型证据包的最多字符数，默认 1200",
    )
    parser.add_argument("--web-results", type=int, default=8, help="Tavily/Brave 每个检索式最多返回多少项，默认 8")
    parser.add_argument(
        "--mce-max-products",
        type=int,
        default=DEFAULT_MCE_MAX_PRODUCTS,
        help="MCE 定向查询的可选软上限；0 表示不截断、处理全部产品",
    )
    parser.add_argument(
        "--include-recent-events",
        action="store_true",
        help="额外执行名称+近期事件查询；结果按间接候选保存，默认不进入大模型证据包",
    )
    parser.add_argument("--pack-items", type=int, default=14, help="每个产品写入大模型证据包的最多证据数，默认 14")
    parser.add_argument("--snippet-chars", type=int, default=700, help="证据包内每条摘要的最多字符数，默认 700")
    parser.add_argument("--tavily-api-key", default=os.getenv("TAVILY_API_KEY", ""), help="Tavily API key；推荐通过环境变量提供")
    parser.add_argument("--brave-api-key", default=os.getenv("BRAVE_SEARCH_API_KEY", ""), help="Brave Search API key；推荐通过环境变量提供")
    parser.add_argument("--ncbi-api-key", default=os.getenv("NCBI_API_KEY", ""), help="可选 NCBI API key；推荐通过环境变量提供")
    parser.add_argument("--ncbi-email", default=os.getenv("NCBI_EMAIL", "research@example.com"), help="NCBI E-utilities 联系邮箱")
    parser.add_argument("--user-agent", default=os.getenv("INQUIRY_RESEARCH_USER_AGENT", "InquiryEvidenceCollector/0.1 (research@example.com)"))
    parser.add_argument("--version", action="version", version=SCRIPT_VERSION)
    args = parser.parse_args(argv)
    args.cas = clean_text(args.cas)
    input_choices = sum(bool(value) for value in (args.input, args.cas, args.resume_dir))
    if input_choices != 1:
        parser.error("必须且只能选择一种输入：文件路径、--cas 或 --resume-dir")
    if args.input and not args.input.exists():
        parser.error(f"输入文件不存在：{args.input}")
    if args.resume_dir and not args.resume_dir.is_dir():
        parser.error(f"续跑目录不存在：{args.resume_dir}")
    if args.cas and not cas_is_valid(args.cas):
        parser.error(f"CAS 格式或校验位无效：{args.cas}")
    if args.mce_max_products < 0:
        parser.error("--mce-max-products 不能为负数；0 表示处理全部产品")
    if args.patent_pages < 0:
        parser.error("--patent-pages 不能为负数")
    if args.patent_candidate_cap < 1:
        parser.error("--patent-candidate-cap 必须至少为 1")
    if args.patent_snippet_chars < 300:
        parser.error("--patent-snippet-chars 必须至少为 300")
    enabled = {item.strip().casefold() for item in args.sources.split(",") if item.strip()}
    unknown = enabled.difference(SUPPORTED_SOURCES)
    if unknown:
        parser.error(f"未知数据源：{', '.join(sorted(unknown))}")
    args.sources = enabled
    args.tavily_api_key = clean_text(args.tavily_api_key)
    args.brave_api_key = clean_text(args.brave_api_key)
    args.ncbi_api_key = clean_text(args.ncbi_api_key)
    if args.resume_dir and args.output_dir is not None:
        parser.error("--resume-dir 会原地续跑，不能同时使用 --output-dir")
    if args.output_dir is None and not args.resume_dir:
        stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        args.output_dir = Path("outputs") / f"evidence-{stamp}"
    return args


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    args = parse_args(argv)
    started_at = utc_now_iso()
    resume_state: dict[str, Any]
    try:
        if args.resume_dir:
            resume_state = load_resume_state(args.resume_dir)
            products = apply_resume_config(args, resume_state)
            resume_state["status"] = "running"
            resume_state["last_resumed_at"] = started_at
            resume_state["resume_count"] = int(resume_state.get("resume_count") or 0) + 1
            args.resumed_from_checkpoint = True
            args.resume_count = resume_state["resume_count"]
        else:
            products = (
                [ProductInput(row_number=1, bd="", cas=args.cas)]
                if args.cas
                else read_products(args.input, args.sheet, args.bd_column, args.cas_column)
            )
            if args.limit > 0:
                products = products[: args.limit]
            assert args.output_dir is not None
            if (args.output_dir / RESUME_STATE_FILENAME).exists():
                raise CollectionError(
                    f"输出目录已包含续跑状态；请改用 --resume-dir {args.output_dir}"
                )
            resume_state = create_resume_state(products, args, started_at)
            args.resumed_from_checkpoint = False
            args.resume_count = 0
    except CollectionError as exc:
        print(f"输入错误：{exc}", file=sys.stderr)
        return 2
    assert args.output_dir is not None
    args.output_dir.mkdir(parents=True, exist_ok=True)
    records_by_row = load_checkpoint_records(args.output_dir, products)
    args.checkpoint_products_reused = len(records_by_row)
    update_resume_progress(args.output_dir, resume_state, products, records_by_row, status="running")
    pending_products = [
        product for product in products if product.row_number not in records_by_row
    ]
    args.mce_allowed_rows = mce_allowed_product_rows(products, args.mce_max_products)
    print(
        f"读取到 {len(products)} 个产品；已复用检查点 {len(records_by_row)} 个；"
        f"待处理 {len(pending_products)} 个；数据源：{', '.join(sorted(args.sources))}"
    )
    if "tavily" in args.sources and not args.tavily_api_key:
        print("提示：未设置 TAVILY_API_KEY，将跳过 Tavily 开放网页检索。")
    if "mce" in args.sources and not args.tavily_api_key:
        print("提示：未设置 TAVILY_API_KEY，将跳过 MCE 定向域名检索。")
    if "mce" in args.sources and args.mce_max_products > 0 and len(products) > args.mce_max_products:
        print(
            f"安全限额：本次仅对前 {args.mce_max_products} 个产品执行 MCE 定向查询；"
            f"其余 {len(products) - args.mce_max_products} 个将标记为 skipped。"
        )
    if "brave" in args.sources and not args.brave_api_key:
        print("提示：未设置 BRAVE_SEARCH_API_KEY，将跳过开放网页检索。")

    client = HttpClient(
        cache_dir=args.cache_dir,
        ttl_days=args.cache_ttl_days,
        timeout=args.timeout,
        retries=args.retries,
        offline=args.offline,
        user_agent=args.user_agent,
        ncbi_api_key=args.ncbi_api_key or None,
    )
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, args.max_workers)) as executor:
            futures = {
                executor.submit(collect_one, product, client, args.sources, args): product
                for product in pending_products
            }
            for completed, future in enumerate(concurrent.futures.as_completed(futures), start=1):
                product = futures[future]
                try:
                    record = future.result()
                except Exception as exc:
                    record = {
                        "input": dataclasses.asdict(product),
                        "cas_valid": cas_is_valid(product.cas),
                        "identity": {},
                        "source_status": {
                            "collector": {
                                "status": "error",
                                "items": 0,
                                "error": f"{type(exc).__name__}: {exc}",
                            }
                        },
                        "source_data": {
                            "collector": {"error": f"{type(exc).__name__}: {exc}"}
                        },
                        "evidence_items": [],
                    }
                records_by_row[product.row_number] = record
                write_json_atomic(checkpoint_path(args.output_dir, product), record)
                update_resume_progress(args.output_dir, resume_state, products, records_by_row)
                status_counts = defaultdict(int)
                for status in record["source_status"].values():
                    status_counts[status.get("status", "unknown")] += 1
                print(
                    f"[{len(records_by_row)}/{len(products)}] {product.bd or '-'} | {product.cas} | "
                    f"证据 {len(record['evidence_items'])} | ok={status_counts['ok']} "
                    f"error={status_counts['error']} skipped={status_counts['skipped']} | 已写入检查点"
                )
    except BaseException:
        update_resume_progress(
            args.output_dir,
            resume_state,
            products,
            records_by_row,
            status="interrupted",
            final_outputs_ready=False,
        )
        raise
    records = [records_by_row[product.row_number] for product in products]
    write_outputs(records, args.output_dir, args, started_at)
    update_resume_progress(
        args.output_dir,
        resume_state,
        products,
        records_by_row,
        status="complete",
        final_outputs_ready=True,
    )
    errors = sum(1 for record in records for status in record["source_status"].values() if status.get("status") == "error")
    print(f"完成：{args.output_dir.resolve()}")
    print(f"产品 {len(records)}；数据源错误 {errors}；大模型证据包：{(args.output_dir / 'llm_evidence_pack.md').resolve()}")
    return 0 if errors == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
